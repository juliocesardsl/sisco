"""Lógica de verificação de conformidade - refatorada do exemplo.py para Django"""
from typing import Dict, List, Tuple
from functools import lru_cache
from pathlib import Path
import json
import re
import unicodedata
import logging
import sys
import datetime
import math
from io import BytesIO
from types import SimpleNamespace

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)


@lru_cache(maxsize=1)
def _carregar_dados_legislacao_10014():
    """Carrega o JSON com os valores de referência funcional para a rubrica 10014."""
    base_dir = Path(__file__).resolve().parent.parent
    json_path = base_dir / 'biblioteca_legislacao' / 'decreto_40610.json'
    if not json_path.exists():
        return {}
    with open(json_path, encoding='utf-8') as handle:
        return json.load(handle)


def _normalizar_codigo_funcional(texto):
    """Normaliza códigos funcionais como CPC-03, CPC03 ou CPC 03 para o formato CPC-03."""
    if texto is None:
        return None
    text = re.sub(r'[^A-Z0-9]+', '', str(texto).strip().upper())
    if not text:
        return None
    match = re.fullmatch(r'([A-Z]+)(\d+)', text)
    if match:
        prefix, suffix = match.groups()
        return f'{prefix}-{int(suffix):02d}'
    return text


def _obter_valor_funcional_json(funcao, campo='remuneracao'):
    """Busca o valor de uma função funcional no JSON de legislação."""
    if not funcao:
        return None

    dados = _carregar_dados_legislacao_10014()
    codigo_alvo = _normalizar_codigo_funcional(funcao)
    if not codigo_alvo:
        return None

    for anexo_key in ('ANEXO_II', 'ANEXO_I'):
        anexo = dados.get(anexo_key, {})
        if not anexo:
            continue
        for codigo, valores in anexo.items():
            codigo_normalizado = _normalizar_codigo_funcional(codigo)
            if codigo_normalizado == codigo_alvo:
                valor = valores.get(campo)
                if valor is not None:
                    return float(valor)
    return None


def _obter_codigo_funcional(row, col_vertical=None, col_horizontal=None):
    """Retorna o código funcional combinando vertical e horizontal quando necessário."""
    def clean_value(value):
        if value is None:
            return None
        text = format_raw_value(value)
        if not text:
            return None
        text = str(text).strip()
        if text.lower() in {'nan', 'none', ''}:
            return None
        return text

    def normalize_function_code(value):
        return _normalizar_codigo_funcional(value)

    def try_col(col):
        if col and col in row:
            return clean_value(row[col])
        return None

    valores = []

    for col in [col_vertical, col_horizontal]:
        valor = try_col(col)
        if valor:
            valores.append(valor)

    if len(valores) >= 2:
        primeiro = normalize_function_code(valores[0])
        segundo = normalize_function_code(valores[1])

        if primeiro and segundo:
            if primeiro == segundo:
                return primeiro

            prefix = re.sub(r'[^A-Z]+', '', primeiro)
            suffix = re.sub(r'[^0-9]+', '', segundo)
            if prefix and suffix:
                try:
                    return f'{prefix}-{int(suffix):02d}'
                except ValueError:
                    return f'{prefix}-{suffix}'

            if re.search(r'\d', primeiro) and re.search(r'[A-Z]', segundo):
                return f'{segundo}-{primeiro}'

            return f'{primeiro}-{segundo}'

    if col_horizontal:
        v = try_col(col_horizontal)
        if v:
            return normalize_function_code(v)
    if col_vertical:
        v = try_col(col_vertical)
        if v:
            return normalize_function_code(v)

    # Fallback: procurar colunas no row com nomes que contenham 'func' + 'horiz'/'hor'
    for col in getattr(row, 'index', row.keys()):
        col_norm = normalize(col)
        if 'func' in col_norm and ('horiz' in col_norm or 'hor' in col_norm):
            valor = clean_value(row[col])
            if valor:
                return valor

    # Fallback2: procurar 'func' + 'vert'
    for col in getattr(row, 'index', row.keys()):
        col_norm = normalize(col)
        if 'func' in col_norm and 'vert' in col_norm:
            valor = clean_value(row[col])
            if valor:
                return valor

    # Último recurso: qualquer coluna contendo 'func'
    for col in getattr(row, 'index', row.keys()):
        col_norm = normalize(col)
        if 'func' in col_norm:
            valor = clean_value(row[col])
            if valor:
                return valor

    return None


def _calcular_valor_rubrica_10014(funcao, frequencia=None):
    """Calcula o valor esperado para a rubrica 10014 considerando frequência e remuneração."""
    valor_remuneracao = _obter_valor_funcional_json(funcao, campo='remuneracao')
    if valor_remuneracao is None:
        return None

    try:
        frequencia_num = float(frequencia) if frequencia is not None else None
    except Exception:
        frequencia_num = None

    if frequencia_num is None or frequencia_num <= 0:
        return round(valor_remuneracao / 0.80, 2)

    valor_por_dia = valor_remuneracao / 30
    valor_com_frequencia = valor_por_dia * frequencia_num
    return round(valor_com_frequencia / 0.80, 2)


def debug_print(msg):
    """Print com flush garantido para aparecer no console Django."""
    try:
        print(msg)
    except UnicodeEncodeError:
        try:
            encoded = msg.encode(sys.stdout.encoding or 'utf-8', errors='replace')
            print(encoded.decode(sys.stdout.encoding or 'utf-8', errors='replace'))
        except Exception:
            pass
    except Exception:
        pass
    try:
        sys.stdout.flush()
    except Exception:
        pass


def flatten_and_clean_columns(df):
    """Limpa headers multilinhas/complexos para strings simples"""
    df.columns = [
        ' '.join(
            str(col).split() if isinstance(col, str) else [str(col)]
        ).strip() if col != '' else f'Unnamed_{i}'
        for i, col in enumerate(df.columns)
    ]
    return df


def _repair_mojibake(s):
    """Tenta corrigir cabeçalhos corrompidos por codificação Latin1/UTF-8."""
    if not isinstance(s, str):
        return s
    if 'Ã' in s or 'Â' in s:
        try:
            return s.encode('latin1').decode('utf-8')
        except Exception:
            return s
    return s

STATUS_DESCRIPTION_MAP = {
    '1': '1 - INCLUÍDO NO MÊS',
    '2': '2 - NORMAL',
    '3': '3 - AFASTADO',
    '4': '4 - DESLIGADO NO MÊS',
    '8': '8 - CEDIDO',
}


def normalize_status_description(value):
    """Mapeia código numérico para descrição de status e mantém texto original se já estiver presente."""
    if value is None:
        return ''

    if isinstance(value, float) and value.is_integer():
        value = str(int(value))
    s = str(value).strip()
    if s == '':
        return ''

    match = re.match(r'^\s*(\d+)\s*(?:-|–)?\s*(.*)$', s)
    if match:
        code = match.group(1)
        if code in STATUS_DESCRIPTION_MAP:
            return STATUS_DESCRIPTION_MAP[code]
        return s

    cleaned = re.sub(r'\.0+$', '', s)
    if cleaned.isdigit() and cleaned in STATUS_DESCRIPTION_MAP:
        return STATUS_DESCRIPTION_MAP[cleaned]

    return s


def extract_status_code(value):
    """Extrai apenas o número do status, preservando texto se não houver número."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    if s == '':
        return ''
    match = re.match(r'^\s*(\d+)\b', s)
    if match:
        return match.group(1)
    return s


def _is_aposentado_status(value):
    """Retorna True se o valor indicar servidor aposentado ou pensionista."""
    if value is None:
        return False
    s = normalize(str(value))
    return 'aposentad' in s or 'pensionista' in s


def _is_servidor_aposentado(row, col_situacao_funcional=None, col_status=None, col_descricao_status=None):
    """Detecta aposentadoria usando colunas de situação funcional e status."""
    if col_situacao_funcional and col_situacao_funcional in row and _is_aposentado_status(row[col_situacao_funcional]):
        return True
    if col_status and col_status in row and _is_aposentado_status(row[col_status]):
        return True
    if col_descricao_status and col_descricao_status in row and _is_aposentado_status(row[col_descricao_status]):
        return True
    return False


def normalize(s):
    """Normaliza string removendo acentos, espaços extras"""
    if not isinstance(s, str):
        return ""
    s = _repair_mojibake(s)
    s = s.strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join([c for c in s if not unicodedata.combining(c)])
    return s


def format_raw_value(value):
    """Preserva valores brutos do Excel sem modificar o formato sempre que possível."""
    if value is None:
        return ''
    try:
        import pandas as pd
        if pd.isna(value):
            return ''
    except Exception:
        pass

    if isinstance(value, datetime.datetime):
        if value.time() == datetime.time(0, 0):
            return value.strftime('%d/%m/%Y')
        return value.strftime('%d/%m/%Y %H:%M:%S')

    if isinstance(value, datetime.date):
        return value.strftime('%d/%m/%Y')

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _extrair_referencia_atual(file_obj):
    """Extrai o mês/ano do arquivo com base no nome do arquivo."""
    name = getattr(file_obj, 'name', '') or ''
    name = name.split('/')[-1].split('\\')[-1]

    # Aceita formatos como 042026, 04_2026, 04-2026, 2026-04, 2026_04, 04/2026
    patterns = [
        (r'(?<!\d)(\d{1,2})[^\d]?(\d{4})(?!\d)', True),
        (r'(?<!\d)(\d{4})[^\d]?(\d{1,2})(?!\d)', False),
    ]

    for pattern, month_first in patterns:
        for match in re.finditer(pattern, name):
            group1, group2 = match.groups()
            mes, ano = (group1, group2) if month_first else (group2, group1)
            try:
                mes_int = int(mes)
                if 1 <= mes_int <= 12:
                    return f"{mes_int:02d}{ano}"
            except ValueError:
                continue

    return ''


def _column_tokens(col):
    """Converte o nome da coluna em tokens normalizados para comparação flexível."""
    normalized = normalize(col)
    return set(token for token in re.sub(r'[^a-z0-9]+', ' ', normalized).split() if token)


COLUMN_ALIASES = {
    'ano': {'ano', 'anorefer'},
    'referencia': {'referencia', 'refer', 'ref', 'anorefer'},
    'ref': {'ref', 'refer', 'referencia', 'anorefer'},
    'mes': {'mes', 'mesrefer', 'mês'},
    'carga': {'carga', 'cargahoraria', 'ch'},
    'horaria': {'horaria', 'horario', 'cargahoraria'},
    'valor': {'valor', 'vl', 'vlrubrica', 'vl_rubrica'},
    'nome': {'nome', 'nomecompleto', 'nomedo'},
    'servidor': {'servidor', 'nomecompleto', 'nomedo'},
    'descricao': {'descricao', 'descrio', 'desc'},
    'desc': {'desc', 'descricao', 'descrio', 'rubrica'},
    'prov': {'prov', 'rubrica', 'desc', 'provdesc', 'idprovdesc', 'cdprovdesc'},
    'empresa': {'empresa', 'emp'},
    'matricula': {'matricula', 'matr'},
    'orgao': {'orgao', 'orga', 'org', 'unidade', 'lotacao', 'lotac'},
    'cpf': {'cpf', 'documento', 'cpfdocumento'},
    'cargo': {'cargo'},
    'funcional': {'funcional', 'funcao'},
    'status': {'status', 'dcstatus'},
    'frequencia': {'frequencia'},
    'situacao': {'situacao', 'situacaofuncional'},
    'data': {'data', 'datadevigencia'},
    'vigencia': {'vigencia', 'vigencia', 'vigencia'},
}


def _matches_word(col_norm, col_clean, col_tokens, word):
    aliases = COLUMN_ALIASES.get(word, {word})
    return (
        word in col_norm
        or word in col_clean
        or any(alias in col_tokens for alias in aliases)
    )


def _find_carga_horaria_column(df):
    """Prioriza a coluna principal de carga horária e evita retornar a coluna secundária como principal."""
    preferred_names = [
        'CARGA_HORARIA',
        'CARGA HORARIA',
        'CARGA_HORÁRIA',
        'CARGA HORÁRIA',
        'CARGA HORARIA PRINCIPAL',
        'CARGA_HORARIA_PRINCIPAL',
        'CARGA HORÁRIA PRINCIPAL',
    ]
    for name in preferred_names:
        if name in df.columns:
            return name

    candidates = []
    for col in df.columns:
        col_norm = normalize(col)
        if 'carga' not in col_norm or ('horaria' not in col_norm and 'horario' not in col_norm):
            continue
        if 'secundaria' in col_norm or 'secundaria' in col_norm or 'secundário' in col_norm or 'secundaria' in col_norm:
            continue
        candidates.append(col)

    if candidates:
        return candidates[0]

    return find_by_words(df, ['carga', 'horaria'])


def _is_empresa_like_column(col):
    col_norm = normalize(col)
    return any(term in col_norm for term in ['empresa', 'emp'])


def _find_nome_servidor_column(df):
    # Prefer columns que explicitem servidor, mas evite colunas de empresa.
    for col in df.columns:
        col_norm = normalize(col)
        if 'servidor' in col_norm and not _is_empresa_like_column(col):
            return col

    # Em seguida, prefira qualquer coluna de nome sem referência a empresa.
    nome_candidates = []
    for col in df.columns:
        col_norm = normalize(col)
        if 'nome' in col_norm and not _is_empresa_like_column(col):
            nome_candidates.append(col)

    if nome_candidates:
        # Prioriza colunas que mencionem 'servidor' junto com 'nome'.
        nome_candidates.sort(key=lambda c: ('servidor' in normalize(c), 'nome' in normalize(c)), reverse=True)
        return nome_candidates[0]

    # Último recurso: relaxa a busca por servidor/nome usando find_by_words,
    # mas ainda evitando colunas de empresa quando possível.
    candidato = find_by_words(df, ['nome', 'servidor'])
    if candidato and not _is_empresa_like_column(candidato):
        return candidato
    candidato = find_by_words(df, ['nome'])
    if candidato and not _is_empresa_like_column(candidato):
        return candidato
    return candidato


def _score_rubrica_code_column(series):
    values = series.dropna().astype(str).str.strip()
    if values.empty:
        return -100

    int_like = values.str.fullmatch(r'\d+').mean()
    decimal_like = values.str.fullmatch(r'\d+[.,]\d{2}').mean()
    money_like = values.str.contains(r'[Rr]\$') | values.str.contains(r'[€£]') | values.str.fullmatch(r'\d+[.,]\d{2}')
    money_like = money_like.mean()

    # Prefer pure numeric codes and penalize values that look like currency/amounts.
    return int_like * 10 - decimal_like * 4 - money_like * 8


def _is_five_digit_column(series):
    values = series.dropna().astype(str).str.strip()
    if values.empty:
        return False
    exact_five = values.str.fullmatch(r'\d{5}')
    return exact_five.mean() >= 0.5


def _is_amount_like_column_name(col):
    if not isinstance(col, str):
        return False
    normalized = normalize(col)
    return any(token in normalized for token in [
        'valor', 'vl', 'vlor', 'venc', 'pagamento', 'total', 'recebido', 'salario', 'receb', 'saldo'
    ])


def _find_rubrica_column(df, debug=False):
    col = _select_rubrica_code_column(df)
    if col:
        return col

    if debug:
        debug_print('  [DEBUG] Procurando coluna de rubrica sem nomes preferenciais...')

    for finder in [
        lambda: find_by_words(df, ['rubrica'], debug=debug),
        lambda: find_by_words(df, ['prov', 'desc'], debug=debug),
    ]:
        candidate = finder()
        if candidate and not _is_amount_like_column_name(candidate):
            return candidate

    for col in df.columns:
        if _is_amount_like_column_name(col):
            continue
        col_norm = normalize(col)
        if any(term in col_norm for term in ['rubrica', 'prov', 'desc', 'dc']):
            return col

    return next((col for col in df.columns if _is_five_digit_column(df[col])), None)


def _select_rubrica_code_column(df):
    candidates = [col for col in ['RUBRICA', 'CD_PROV_DESC', 'ID_PROV_DESC', 'PROV/DESC'] if col in df.columns]
    if not candidates:
        candidates = [col for col in df.columns if _is_five_digit_column(df[col])]
    if not candidates:
        return None

    candidates.sort(key=lambda col: ['RUBRICA', 'CD_PROV_DESC', 'ID_PROV_DESC', 'PROV/DESC'].index(col) if col in ['RUBRICA', 'CD_PROV_DESC', 'ID_PROV_DESC', 'PROV/DESC'] else len(['RUBRICA', 'CD_PROV_DESC', 'ID_PROV_DESC', 'PROV/DESC']))
    candidates = sorted(candidates, key=lambda col: _score_rubrica_code_column(df[col]), reverse=True)
    return candidates[0]


def find_by_words(df, words, debug=False):
    """Encontra coluna que contém todas as palavras, aceitando aliases do formato novo."""
    words = [normalize(w) for w in words]

    if debug:
        print(f"  [DEBUG find_by_words] Procurando: {words}")

    # Prioridade: manter compatibilidade com o cabeçalho clássico "PROV/DESC"
    if any(w in ['prov', 'desc'] for w in words):
        if 'PROV/DESC' in df.columns:
            if debug:
                print(f"  [DEBUG] Coluna 'PROV/DESC' encontrada (prioritária)")
            return 'PROV/DESC'

    for col in df.columns:
        col_norm = normalize(col)
        col_clean = ''.join(c for c in col_norm if c.isalnum())
        col_tokens = _column_tokens(col)
        if all(_matches_word(col_norm, col_clean, col_tokens, w) for w in words):
            if debug:
                print(f"  [DEBUG] Match flexível encontrado: '{col}'")
            return col

    if len(words) == 1:
        word = words[0]
        for col in df.columns:
            col_norm = normalize(col)
            col_clean = ''.join(c for c in col_norm if c.isalnum())
            col_tokens = _column_tokens(col)
            if _matches_word(col_norm, col_clean, col_tokens, word):
                if debug:
                    print(f"  [DEBUG] Match por palavra única encontrado: '{col}'")
                return col

    if debug:
        print(f"  [DEBUG] Nenhuma coluna encontrada!")
    return None


def _read_uploaded_excel(file):
    import pandas as pd
    import warnings

    if hasattr(file, 'temporary_file_path'):
        return read_excel_flexible(file.temporary_file_path())

    if hasattr(file, 'read'):
        content = file.read()
        try:
            file.seek(0)
        except Exception:
            pass

        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', UserWarning)
                return pd.read_excel(BytesIO(content), engine='openpyxl')
        except Exception:
            # Fallback to the flexible reader that can detect HTML / engine issues
            return read_excel_flexible(BytesIO(content))

    return read_excel_flexible(file)


def _normalize_value_column(series):
    import pandas as pd
    import numpy as np

    def parse_value(val):
        if pd.isna(val) or val == '' or str(val).strip() == '':
            return 0.0

        # Se já é numérico, retornar diretamente
        if isinstance(val, (int, float)):
            return float(val)

        # Converter para string e limpar espaços
        s = str(val).strip()

        # Remover caracteres de moeda comuns
        s = s.replace('R$', '').replace('$', '').replace('€', '').replace('£', '').strip()

        # Tentar conversão direta primeiro
        try:
            # Substituir vírgula por ponto se for formato brasileiro simples
            if ',' in s and s.count(',') == 1 and '.' not in s:
                return float(s.replace(',', '.'))
            # Formato americano
            elif '.' in s and ',' not in s:
                return float(s)
            # Formato com milhares brasileiro: 1.234,56
            elif ',' in s and '.' in s:
                # Verificar se é formato brasileiro com separador de milhares
                parts = s.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    # Remover pontos dos milhares
                    integer_part = parts[0].replace('.', '')
                    return float(f"{integer_part}.{parts[1]}")
        except ValueError:
            pass

        # Método alternativo: extrair apenas dígitos e pontuar
        digits = ''.join(c for c in s if c.isdigit())
        if not digits:
            return 0.0

        # Se tem exatamente 2 dígitos, assumir que é centavos: 00 -> 0.00
        if len(digits) == 2:
            return float(f"0.{digits}")
        # Se tem mais, assumir que os últimos 2 são centavos
        elif len(digits) > 2:
            return float(f"{digits[:-2]}.{digits[-2:]}")
        else:
            return float(digits)

    return series.apply(parse_value)


def _parse_extrator_for_rubrica(file, rubrica):
    import pandas as pd

    df_extrator = _read_uploaded_excel(file)
    df_extrator = flatten_and_clean_columns(df_extrator)

    col_prov = _find_rubrica_column(df_extrator)

    col_empresa = (find_by_words(df_extrator, ['empresa']) or 
                   find_by_words(df_extrator, ['cod', 'empresa']) or 
                   find_by_words(df_extrator, ['codigo', 'empresa']) or
                   find_by_words(df_extrator, ['cod_empresa']) or
                   find_by_words(df_extrator, ['código']))
    col_cargo = (find_by_words(df_extrator, ['cargo']) or
                 find_by_words(df_extrator, ['funcao']) or
                 find_by_words(df_extrator, ['cbo']) or
                 find_by_words(df_extrator, ['referencia']))
    if not col_cargo:
        for col in df_extrator.columns:
            col_norm = normalize(col)
            if any(term in col_norm for term in ['cargo', 'funcao', 'cbo', 'referencia', 'cargo']):
                col_cargo = col
                break

    col_matricula = (find_by_words(df_extrator, ['matricula']) or
                     find_by_words(df_extrator, ['matr']) or
                     find_by_words(df_extrator, ['mat']))
    if not col_matricula:
        for col in df_extrator.columns:
            col_norm = normalize(col)
            if any(term in col_norm for term in ['matricula', 'matr', 'mat']):
                col_matricula = col
                break

    col_valor_extrator = next((col for col in ['VL_RUBRICA', 'VLRUBRICA', 'VALOR'] if col in df_extrator.columns), None)
    if not col_valor_extrator:
        col_valor_extrator = find_by_words(df_extrator, ['valor'])
    if not col_valor_extrator:
        for col in df_extrator.columns:
            col_norm = normalize(col)
            if 'valor' in col_norm:
                col_valor_extrator = col
                break

    col_rubrica = _select_rubrica_code_column(df_extrator)

    if not col_prov or not col_empresa or not col_cargo or not col_matricula or not col_valor_extrator:
        faltantes = []
        if not col_prov:
            faltantes.append('PROV/DESC')
        if not col_empresa:
            faltantes.append('EMPRESA')
        if not col_cargo:
            faltantes.append('CARGO')
        if not col_matricula:
            faltantes.append('MATRICULA')
        if not col_valor_extrator:
            faltantes.append('VALOR')
        raise Exception(f"Colunas obrigatórias não encontradas no EXTRATOR: {', '.join(faltantes)}\nColunas disponíveis: {df_extrator.columns.tolist()}")

    df_extrator[col_valor_extrator] = _normalize_value_column(df_extrator[col_valor_extrator])

    # Tentar diferentes formas de filtrar a rubrica
    rubrica_str = str(rubrica).strip()
    filtro_col = col_rubrica if rubrica_str.isdigit() and col_rubrica else col_prov

    # Primeiro tentar filtro exato
    filtro = df_extrator[filtro_col].astype(str).str.strip() == rubrica_str
    df_filtrado = df_extrator[filtro].copy()

    # Se não encontrou, tentar filtro por contém
    if df_filtrado.empty:
        filtro = df_extrator[filtro_col].astype(str).str.contains(rubrica_str, case=False, na=False)
        df_filtrado = df_extrator[filtro].copy()

    # Se ainda não encontrou, tentar apenas com números (código da rubrica)
    if df_filtrado.empty and rubrica_str.isdigit():
        filtro = df_extrator[col_prov].astype(str).str.contains(r'\b' + rubrica_str + r'\b', case=False, na=False)
        df_filtrado = df_extrator[filtro].copy()

    if df_filtrado.empty:
        raise Exception(f"Rubrica '{rubrica_str}' não encontrada no arquivo de extrator informado.\nValores disponíveis na coluna {col_prov}: {df_extrator[col_prov].unique().tolist()[:20]}")

    group_cols = [col_empresa, col_matricula]
    resumo = df_filtrado.groupby(group_cols, dropna=False, as_index=False)[col_valor_extrator].mean()
    resumo = resumo.rename(columns={col_empresa: 'empresa', col_matricula: 'matricula', col_valor_extrator: 'valor'})
    resumo['empresa'] = resumo['empresa'].astype(str).str.strip().str.zfill(3)
    resumo['matricula'] = resumo['matricula'].astype(str).str.strip()

    return resumo


def _listar_rubricas_no_extrator(file):
    """Retorna as rubricas únicas encontradas em um arquivo de extrator."""
    df_extrator = _read_uploaded_excel(file)
    df_extrator = flatten_and_clean_columns(df_extrator)

    col_prov = _find_rubrica_column(df_extrator)

    if not col_prov:
        raise Exception('Não foi possível identificar a coluna de rubrica no arquivo de extrator.')

    valores = df_extrator[col_prov].dropna().astype(str).str.strip()
    rubricas = []
    vistos = set()
    for valor in valores:
        if not valor or valor in vistos:
            continue
        vistos.add(valor)
        rubricas.append(valor)

    return rubricas


def _parse_extrator_for_rubrica_detailed(file, rubrica):
    import pandas as pd

    df_extrator = _read_uploaded_excel(file)
    df_extrator = flatten_and_clean_columns(df_extrator)

    col_prov = _find_rubrica_column(df_extrator)

    col_rubrica = _select_rubrica_code_column(df_extrator)
    if not col_rubrica:
        col_rubrica = find_by_words(df_extrator, ['rubrica'])

    col_dc_rubrica = next((col for col in ['DC_RUBRICA', 'DESCRICAO_RUBRICA', 'DESCRICAO DA RUBRICA', 'DESC_RUBRICA'] if col in df_extrator.columns), None)
    if not col_dc_rubrica:
        col_dc_rubrica = find_by_words(df_extrator, ['dc', 'rubrica']) or find_by_words(df_extrator, ['descricao', 'rubrica'])

    col_empresa = (find_by_words(df_extrator, ['empresa']) or 
                   find_by_words(df_extrator, ['cod', 'empresa']) or 
                   find_by_words(df_extrator, ['codigo', 'empresa']) or
                   find_by_words(df_extrator, ['cod_empresa']) or
                   find_by_words(df_extrator, ['código']))
    col_matricula = find_by_words(df_extrator, ['matricula', 'matrícula'])
    col_valor_extrator = next((col for col in ['VL_RUBRICA', 'VLRUBRICA', 'VALOR'] if col in df_extrator.columns), None)
    if not col_valor_extrator:
        col_valor_extrator = find_by_words(df_extrator, ['valor'])

    # Additional detail columns
    col_nome_servidor = _find_nome_servidor_column(df_extrator)
    col_cpf = find_by_words(df_extrator, ['cpf']) or find_by_words(df_extrator, ['documento']) or find_by_words(df_extrator, ['cpf/documento'])
    col_situacao_funcional = find_by_words(df_extrator, ['situacao', 'funcional']) or find_by_words(df_extrator, ['situação', 'funcional'])
    # Buscar coluna de status (código) priorizando nomes exatos e evitando descrições
    col_status = None
    for col in df_extrator.columns:
        col_norm = normalize(col)
        if col_norm in ['status', 'codstatus', 'codigo_status', 'código_status']:
            col_status = col
            break
    if not col_status:
        # fallback para busca flexível, mas evitando colunas de descrição
        for col in df_extrator.columns:
            col_norm = normalize(col)
            if 'status' in col_norm and 'desc' not in col_norm and 'descri' not in col_norm:
                col_status = col
                break
    if not col_status:
        col_status = find_by_words(df_extrator, ['status'])

    # Buscar coluna de descrição do status
    col_descricao_status = None
    for col in df_extrator.columns:
        col_norm = normalize(col)
        if ('desc' in col_norm or 'descri' in col_norm) and 'status' in col_norm:
            col_descricao_status = col
            break
    if not col_descricao_status:
        col_descricao_status = find_by_words(df_extrator, ['descricao', 'status']) or find_by_words(df_extrator, ['dc_status'])
    col_cargo = find_by_words(df_extrator, ['cargo'])
    col_descricao_cargo = next((col for col in ['DC_CARREIRA', 'DESCRICAO_CARREIRA', 'DESCRIÇÃO_CARREIRA'] if col in df_extrator.columns), None)
    col_descricao_cargo = col_descricao_cargo or find_by_words(df_extrator, ['descricao', 'cargo']) or find_by_words(df_extrator, ['descrição', 'cargo'])
    col_data_admissao = find_by_words(df_extrator, ['data', 'admissao']) or find_by_words(df_extrator, ['data', 'admissão'])
    col_data_ingresso_ref_salarial = find_by_words(df_extrator, ['data', 'ingresso', 'referencia', 'salarial']) or find_by_words(df_extrator, ['data', 'ingresso', 'ref', 'salarial'])
    col_data_afastamento = find_by_words(df_extrator, ['data', 'afastamento'])
    col_motivo_afastamento = find_by_words(df_extrator, ['motivo', 'afastamento'])
    col_motivo_desligamento = find_by_words(df_extrator, ['motivo', 'desligamento']) or find_by_words(df_extrator, ['motivo', 'demissão'])
    col_carga_horaria = _find_carga_horaria_column(df_extrator)
    col_carga_horaria_secundaria = find_by_words(df_extrator, ['carga', 'horaria', 'secundaria']) or find_by_words(df_extrator, ['carga', 'horária', 'secundária'])
    col_ref_vertical = find_by_words(df_extrator, ['ref', 'salarial', 'vertical'])
    col_ref_horizontal = find_by_words(df_extrator, ['ref', 'salarial', 'horizontal'])
    col_prov_desc = find_by_words(df_extrator, ['prov', 'desc'])
    col_valor_mes_anterior = (find_by_words(df_extrator, ['valor', 'mes', 'anterior']) or 
                                        find_by_words(df_extrator, ['valor', 'mês', 'anterior']) or
                                        find_by_words(df_extrator, ['valor', 'anterior']))
    col_valor_mes_atual = (find_by_words(df_extrator, ['valor', 'mes', 'atual']) or 
                                    find_by_words(df_extrator, ['valor', 'mês', 'atual']) or
                                    find_by_words(df_extrator, ['valor', 'atual']))
    col_valor_vencimento = (find_by_words(df_extrator, ['valor', 'vencimento']) or 
                            find_by_words(df_extrator, ['vencimento']))
    col_valor_total_recebido = (find_by_words(df_extrator, ['valor', 'total', 'recebido']) or
                                find_by_words(df_extrator, ['valor', 'recebido']) or
                                find_by_words(df_extrator, ['total', 'recebido']) or
                                find_by_words(df_extrator, ['recebido']))
    col_frequencia = (find_by_words(df_extrator, ['frequencia']) or 
                      find_by_words(df_extrator, ['frequência']))

    if not col_prov or not col_empresa or not col_matricula or not col_valor_extrator:
        faltantes = []
        if not col_prov:
            faltantes.append('PROV/DESC')
        if not col_empresa:
            faltantes.append('EMPRESA')
        if not col_matricula:
            faltantes.append('MATRICULA')
        if not col_valor_extrator:
            faltantes.append('VALOR')
        raise Exception(f"Colunas obrigatórias não encontradas no EXTRATOR: {', '.join(faltantes)}")

    df_extrator[col_valor_extrator] = _normalize_value_column(df_extrator[col_valor_extrator])

    rubrica_str = str(rubrica).strip()
    filtro_col = col_rubrica if rubrica_str.isdigit() and col_rubrica else col_prov
    filtro = df_extrator[filtro_col].astype(str).str.strip() == rubrica_str
    df_filtrado = df_extrator[filtro].copy()
    if df_filtrado.empty:
        df_filtrado = df_extrator[df_extrator[filtro_col].astype(str).str.contains(rubrica_str, case=False, na=False)].copy()
    if df_filtrado.empty:
        raise Exception(f"Rubrica '{rubrica_str}' não encontrada no arquivo de extrator informado.")

    def make_column(col_name, default=''):
        if col_name and col_name in df_filtrado.columns:
            return df_filtrado[col_name].apply(format_raw_value)
        return pd.Series([default] * len(df_filtrado), index=df_filtrado.index)

    df_filtrado['empresa'] = df_filtrado[col_empresa].astype(str).str.strip().replace('.0', '', regex=False)
    df_filtrado['matricula'] = df_filtrado[col_matricula].astype(str).str.strip().replace('.0', '', regex=False)
    df_filtrado['rubrica'] = make_column(col_rubrica or col_prov)
    df_filtrado['dc_rubrica'] = make_column(col_dc_rubrica)
    df_filtrado['valor'] = df_filtrado[col_valor_extrator]
    df_filtrado['nome_servidor'] = make_column(col_nome_servidor)
    df_filtrado['cpf'] = make_column(col_cpf)
    df_filtrado['situacao_funcional'] = make_column(col_situacao_funcional)
    df_filtrado['status_servidor'] = make_column(col_status).apply(extract_status_code)
    df_filtrado['descricao_status'] = make_column(col_descricao_status).apply(normalize_status_description)
    df_filtrado['cargo'] = make_column(col_cargo)
    df_filtrado['descricao_cargo'] = make_column(col_descricao_cargo)
    df_filtrado['descricao_status'] = df_filtrado['descricao_status'].where(
        df_filtrado['descricao_status'] != '',
        df_filtrado['status_servidor'].apply(normalize_status_description)
    )
    df_filtrado['data_admissao'] = make_column(col_data_admissao)
    df_filtrado['data_ingresso_ref_salarial'] = make_column(col_data_ingresso_ref_salarial)
    df_filtrado['data_afastamento'] = make_column(col_data_afastamento)
    df_filtrado['motivo_afastamento'] = make_column(col_motivo_afastamento)
    df_filtrado['motivo_desligamento'] = make_column(col_motivo_desligamento)
    df_filtrado['carga_horaria'] = make_column(col_carga_horaria)
    df_filtrado['carga_horaria_secundaria'] = make_column(col_carga_horaria_secundaria)
    df_filtrado['ref_vertical'] = make_column(col_ref_vertical)
    df_filtrado['ref_horizontal'] = make_column(col_ref_horizontal)
    df_filtrado['prov_desc'] = make_column(col_prov)
    df_filtrado['valor_mes_anterior'] = make_column(col_valor_mes_anterior)
    df_filtrado['valor_mes_atual'] = make_column(col_valor_mes_atual)
    df_filtrado['valor_vencimento'] = make_column(col_valor_vencimento)
    df_filtrado['valor_total_recebido'] = make_column(col_valor_total_recebido)
    df_filtrado['frequencia'] = make_column(col_frequencia)

    campos_agregacao = {
        'valor': 'mean',
        'nome_servidor': 'first',
        'cpf': 'first',
        'situacao_funcional': 'first',
        'status_servidor': 'first',
        'descricao_status': 'first',
        'cargo': 'first',
        'descricao_cargo': 'first',
        'rubrica': 'first',
        'dc_rubrica': 'first',
        'data_admissao': 'first',
        'data_ingresso_ref_salarial': 'first',
        'data_afastamento': 'first',
        'motivo_afastamento': 'first',
        'motivo_desligamento': 'first',
        'carga_horaria': 'first',
        'carga_horaria_secundaria': 'first',
        'ref_vertical': 'first',
        'ref_horizontal': 'first',
        'prov_desc': 'first',
        'valor_mes_anterior': 'first',
        'valor_mes_atual': 'first',
        'valor_vencimento': 'first',
        'valor_total_recebido': 'first',
        'frequencia': 'first',
    }

    resumo = df_filtrado.groupby(['empresa', 'matricula'], dropna=False, as_index=False).agg(campos_agregacao)
    resumo['empresa'] = resumo['empresa'].astype(str).str.strip().str.zfill(3)
    resumo['matricula'] = resumo['matricula'].astype(str).str.strip()
    return resumo


def determinar_status_comparacao(valor_anterior, valor_atual):
    """Define o status de comparação para a tela de comparativo."""
    valor_anterior = float(valor_anterior or 0)
    valor_atual = float(valor_atual or 0)

    if valor_anterior == 0 and valor_atual > 0:
        return 'novo'
    if valor_atual == 0 and valor_anterior > 0:
        return 'removido'
    if valor_atual > valor_anterior:
        return 'aumento'
    if valor_atual < valor_anterior:
        return 'reducao'
    return 'sem-variacao'


def calcular_status_variacao(valor_anterior, valor_atual):
    """Retorna o rótulo textual da variação entre dois valores."""
    valor_anterior = float(valor_anterior or 0)
    valor_atual = float(valor_atual or 0)

    if valor_anterior == 0 and valor_atual == 0:
        return ''
    if valor_anterior == 0 and valor_atual > 0:
        return 'Sem pagamento no mês anterior'
    if valor_atual == 0 and valor_anterior > 0:
        return 'Sem pagamento no mês atual'
    if valor_atual > valor_anterior:
        return 'Houve aumento'
    if valor_atual < valor_anterior:
        return 'Houve redução'
    return 'Não houve variação'


def _parse_extrator_para_comparacao(file, rubricas=None):
    """Carrega um extrator uma vez e agrega por empresa/matrícula/rubrica."""
    import pandas as pd

    df_extrator = _read_uploaded_excel(file)
    df_extrator = flatten_and_clean_columns(df_extrator)

    col_prov = _find_rubrica_column(df_extrator)

    col_rubrica = _select_rubrica_code_column(df_extrator)
    if not col_rubrica:
        col_rubrica = find_by_words(df_extrator, ['rubrica'])

    col_dc_rubrica = next((col for col in ['DC_RUBRICA', 'DESCRICAO_RUBRICA', 'DESCRICAO DA RUBRICA', 'DESC_RUBRICA'] if col in df_extrator.columns), None)
    if not col_dc_rubrica:
        col_dc_rubrica = find_by_words(df_extrator, ['dc', 'rubrica']) or find_by_words(df_extrator, ['descricao', 'rubrica'])

    col_empresa = (find_by_words(df_extrator, ['empresa']) or find_by_words(df_extrator, ['cod', 'empresa']) or find_by_words(df_extrator, ['codigo', 'empresa']) or find_by_words(df_extrator, ['cod_empresa']) or find_by_words(df_extrator, ['código']))
    col_matricula = find_by_words(df_extrator, ['matricula', 'matrícula'])
    col_valor_extrator = next((col for col in ['VL_RUBRICA', 'VLRUBRICA', 'VALOR'] if col in df_extrator.columns), None)
    if not col_valor_extrator:
        col_valor_extrator = find_by_words(df_extrator, ['valor'])

    col_nome_servidor = _find_nome_servidor_column(df_extrator)
    col_cpf = find_by_words(df_extrator, ['cpf']) or find_by_words(df_extrator, ['documento']) or find_by_words(df_extrator, ['cpf/documento'])
    col_situacao_funcional = find_by_words(df_extrator, ['situacao', 'funcional']) or find_by_words(df_extrator, ['situação', 'funcional'])
    col_status = None
    for col in df_extrator.columns:
        col_norm = normalize(col)
        if col_norm in ['status', 'codstatus', 'codigo_status', 'código_status']:
            col_status = col
            break
    if not col_status:
        for col in df_extrator.columns:
            col_norm = normalize(col)
            if 'status' in col_norm and 'desc' not in col_norm and 'descri' not in col_norm:
                col_status = col
                break
    if not col_status:
        col_status = find_by_words(df_extrator, ['status'])

    col_descricao_status = None
    for col in df_extrator.columns:
        col_norm = normalize(col)
        if ('desc' in col_norm or 'descri' in col_norm) and 'status' in col_norm:
            col_descricao_status = col
            break
    if not col_descricao_status:
        col_descricao_status = find_by_words(df_extrator, ['descricao', 'status']) or find_by_words(df_extrator, ['dc_status'])

    col_cargo = find_by_words(df_extrator, ['cargo'])
    col_descricao_cargo = next((col for col in ['DC_CATEGORIA', 'DESCRICAO_CATEGORIA', 'DESCRIÇÃO_CATEGORIA'] if col in df_extrator.columns), None)
    col_descricao_cargo = col_descricao_cargo or find_by_words(df_extrator, ['descricao', 'cargo']) or find_by_words(df_extrator, ['descrição', 'cargo'])
    col_nr_parcela_inicial = next((col for col in ['NR_PARCELA_INICIAL', 'NUM_PARCELA_INICIAL'] if col in df_extrator.columns), None)
    col_prazo_parcela = next((col for col in ['PRAZO_PARCELA', 'PRAZO_DA_PARCELA'] if col in df_extrator.columns), None)
    col_data_admissao = find_by_words(df_extrator, ['data', 'admissao']) or find_by_words(df_extrator, ['data', 'admissão'])
    col_data_ingresso_ref_salarial = find_by_words(df_extrator, ['data', 'ingresso', 'referencia', 'salarial']) or find_by_words(df_extrator, ['data', 'ingresso', 'ref', 'salarial'])
    col_data_afastamento = find_by_words(df_extrator, ['data', 'afastamento'])
    col_motivo_afastamento = find_by_words(df_extrator, ['motivo', 'afastamento'])
    col_motivo_desligamento = find_by_words(df_extrator, ['motivo', 'desligamento']) or find_by_words(df_extrator, ['motivo', 'demissão'])
    col_carga_horaria = _find_carga_horaria_column(df_extrator)
    col_carga_horaria_secundaria = find_by_words(df_extrator, ['carga', 'horaria', 'secundaria']) or find_by_words(df_extrator, ['carga', 'horária', 'secundária'])
    col_ref_vertical = find_by_words(df_extrator, ['ref', 'salarial', 'vertical'])
    col_ref_horizontal = find_by_words(df_extrator, ['ref', 'salarial', 'horizontal'])
    col_prov_desc = find_by_words(df_extrator, ['prov', 'desc'])
    col_valor_mes_anterior = (find_by_words(df_extrator, ['valor', 'mes', 'anterior']) or find_by_words(df_extrator, ['valor', 'mês', 'anterior']) or find_by_words(df_extrator, ['valor', 'anterior']))
    col_valor_mes_atual = (find_by_words(df_extrator, ['valor', 'mes', 'atual']) or find_by_words(df_extrator, ['valor', 'mês', 'atual']) or find_by_words(df_extrator, ['valor', 'atual']))
    col_valor_vencimento = find_by_words(df_extrator, ['valor', 'vencimento']) or find_by_words(df_extrator, ['vencimento'])
    col_valor_total_recebido = (find_by_words(df_extrator, ['valor', 'total', 'recebido']) or find_by_words(df_extrator, ['valor', 'recebido']) or find_by_words(df_extrator, ['total', 'recebido']) or find_by_words(df_extrator, ['recebido']))
    col_frequencia = find_by_words(df_extrator, ['frequencia']) or find_by_words(df_extrator, ['frequência'])
    col_versao = find_by_words(df_extrator, ['versao', 'versão']) or find_by_words(df_extrator, ['versao'])
    col_ano_referencia = find_by_words(df_extrator, ['ano', 'referencia']) or find_by_words(df_extrator, ['ano', 'referência'])
    col_mes_referencia = find_by_words(df_extrator, ['mes', 'referencia']) or find_by_words(df_extrator, ['mês', 'referência']) or find_by_words(df_extrator, ['mes', 'referência']) or find_by_words(df_extrator, ['mês', 'referencia'])

    if not col_prov or not col_empresa or not col_matricula or not col_valor_extrator:
        raise Exception('Colunas obrigatórias não encontradas no EXTRATOR para comparação rápida.')

    df_extrator[col_valor_extrator] = _normalize_value_column(df_extrator[col_valor_extrator])

    def make_column(col_name, default=''):
        if col_name and col_name in df_extrator.columns:
            return df_extrator[col_name].apply(format_raw_value)
        return pd.Series([default] * len(df_extrator), index=df_extrator.index)

    df_extrator['empresa'] = df_extrator[col_empresa].astype(str).str.strip().replace('.0', '', regex=False)
    df_extrator['matricula'] = df_extrator[col_matricula].astype(str).str.strip().replace('.0', '', regex=False)
    df_extrator['rubrica'] = make_column(col_rubrica or col_prov)
    df_extrator['dc_rubrica'] = make_column(col_dc_rubrica)
    df_extrator['valor'] = df_extrator[col_valor_extrator]
    df_extrator['nome_servidor'] = make_column(col_nome_servidor)
    df_extrator['cpf'] = make_column(col_cpf)
    df_extrator['situacao_funcional'] = make_column(col_situacao_funcional)
    df_extrator['status_servidor'] = make_column(col_status).apply(extract_status_code)
    df_extrator['descricao_status'] = make_column(col_descricao_status).apply(normalize_status_description)
    df_extrator['cargo'] = make_column(col_cargo)
    df_extrator['descricao_cargo'] = make_column(col_descricao_cargo)
    df_extrator['nr_parcela_inicial'] = make_column(col_nr_parcela_inicial)
    df_extrator['prazo_parcela'] = make_column(col_prazo_parcela)
    df_extrator['data_admissao'] = make_column(col_data_admissao)
    df_extrator['data_ingresso_ref_salarial'] = make_column(col_data_ingresso_ref_salarial)
    df_extrator['data_afastamento'] = make_column(col_data_afastamento)
    df_extrator['motivo_afastamento'] = make_column(col_motivo_afastamento)
    df_extrator['motivo_desligamento'] = make_column(col_motivo_desligamento)
    df_extrator['carga_horaria'] = make_column(col_carga_horaria)
    df_extrator['carga_horaria_secundaria'] = make_column(col_carga_horaria_secundaria)
    df_extrator['ref_vertical'] = make_column(col_ref_vertical)
    df_extrator['ref_horizontal'] = make_column(col_ref_horizontal)
    df_extrator['prov_desc'] = make_column(col_prov)
    df_extrator['valor_mes_anterior'] = make_column(col_valor_mes_anterior)
    df_extrator['valor_mes_atual'] = make_column(col_valor_mes_atual)
    df_extrator['valor_vencimento'] = make_column(col_valor_vencimento)
    df_extrator['valor_total_recebido'] = make_column(col_valor_total_recebido)
    df_extrator['frequencia'] = make_column(col_frequencia)
    df_extrator['versao'] = make_column(col_versao)
    df_extrator['ano_referencia'] = make_column(col_ano_referencia)
    df_extrator['mes_referencia'] = make_column(col_mes_referencia)

    if rubricas:
        rubrica_set = {str(item).strip() for item in rubricas if str(item).strip()}
        if rubrica_set:
            df_extrator = df_extrator[df_extrator['rubrica'].astype(str).str.strip().isin(rubrica_set)].copy()

    campos_agregacao = {
        'valor': 'mean',
        'nome_servidor': 'first', 'cpf': 'first', 'situacao_funcional': 'first', 'status_servidor': 'first',
        'descricao_status': 'first', 'cargo': 'first', 'descricao_cargo': 'first',
        'nr_parcela_inicial': 'first', 'prazo_parcela': 'first', 'rubrica': 'first',
        'dc_rubrica': 'first', 'data_admissao': 'first', 'data_ingresso_ref_salarial': 'first',
        'data_afastamento': 'first', 'motivo_afastamento': 'first', 'motivo_desligamento': 'first',
        'carga_horaria': 'first', 'carga_horaria_secundaria': 'first', 'ref_vertical': 'first',
        'ref_horizontal': 'first', 'prov_desc': 'first', 'valor_mes_anterior': 'first',
        'valor_mes_atual': 'first', 'valor_vencimento': 'first', 'valor_total_recebido': 'first', 'frequencia': 'first',
        'versao': 'first', 'ano_referencia': 'first', 'mes_referencia': 'first',
    }

    resumo = df_extrator.groupby(['empresa', 'matricula', 'rubrica'], dropna=False, as_index=False).agg(campos_agregacao)
    resumo['empresa'] = resumo['empresa'].astype(str).str.strip().str.zfill(3)
    resumo['matricula'] = resumo['matricula'].astype(str).str.strip()
    return resumo


def _construir_referencia(ano, mes):
    """
    Constrói uma referência no formato MMYYYY a partir de ano e mês.
    
    Args:
        ano: String ou int com o ano (ex: "2026", 2026)
        mes: String ou int com o mês (ex: "05", 5)
        
    Returns:
        String no formato MMYYYY (ex: "052026") ou "" se dados inválidos
    """
    if not ano or not mes:
        return ""
    
    try:
        ano_int = int(ano)
        mes_int = int(mes)
        
        if 1 <= mes_int <= 12 and 1900 <= ano_int <= 9999:
            return f"{mes_int:02d}{ano_int}"
    except (ValueError, TypeError):
        pass
    
    return ""


def comparar_extrator_por_mes(file_anterior, file_atual, rubrica):
    # As referências agora serão construídas a partir dos dados ANO REFERENCIA e MES REFERENCIA dos arquivos
    # durante a montagem do item na comparação

    rubricas = []
    if rubrica in (None, '', 'todos', 'todas', 'all'):
        try:
            rubricas = _listar_rubricas_no_extrator(file_anterior)
            if not rubricas:
                rubricas = _listar_rubricas_no_extrator(file_atual)
        except Exception:
            rubricas = []
    else:
        if isinstance(rubrica, (list, tuple, set)):
            rubricas = [str(item).strip() for item in rubrica if str(item).strip()]
        else:
            rubricas = [str(rubrica).strip()]

    if not rubricas:
        return {'erro': 'Nenhuma rubrica foi encontrada nos arquivos informados.'}

    try:
        resumo_anterior = _parse_extrator_para_comparacao(file_anterior, rubricas)
        resumo_atual = _parse_extrator_para_comparacao(file_atual, rubricas)
    except Exception as e:
        return {'erro': str(e)}

    merged = resumo_anterior.merge(
        resumo_atual,
        on=['empresa', 'matricula', 'rubrica'],
        how='outer',
        suffixes=('_anterior', '_atual')
    )
    merged = merged.fillna({
        'valor_anterior': 0,
        'valor_atual': 0,
        'nome_servidor_anterior': '',
        'nome_servidor_atual': '',
        'cpf_anterior': '',
        'cpf_atual': '',
        'situacao_funcional_anterior': '',
        'situacao_funcional_atual': '',
        'status_servidor_anterior': '',
        'status_servidor_atual': '',
        'descricao_status_anterior': '',
        'descricao_status_atual': '',
        'cargo_anterior': '',
        'cargo_atual': '',
        'descricao_cargo_anterior': '',
        'descricao_cargo_atual': '',
        'data_admissao_anterior': '',
        'data_admissao_atual': '',
        'data_ingresso_ref_salarial_anterior': '',
        'data_ingresso_ref_salarial_atual': '',
        'data_afastamento_anterior': '',
        'data_afastamento_atual': '',
        'motivo_afastamento_anterior': '',
        'motivo_afastamento_atual': '',
        'motivo_desligamento_anterior': '',
        'motivo_desligamento_atual': '',
        'carga_horaria_anterior': '',
        'carga_horaria_atual': '',
        'carga_horaria_secundaria_anterior': '',
        'carga_horaria_secundaria_atual': '',
        'ref_vertical_anterior': '',
        'ref_vertical_atual': '',
        'ref_horizontal_anterior': '',
        'ref_horizontal_atual': '',
        'prov_desc_anterior': '',
        'prov_desc_atual': '',
        'valor_mes_anterior_anterior': '',
        'valor_mes_anterior_atual': '',
        'valor_mes_atual_anterior': '',
        'valor_mes_atual_atual': '',
        'valor_vencimento_anterior': '',
        'valor_vencimento_atual': '',
        'valor_total_recebido_anterior': '',
        'valor_total_recebido_atual': '',
        'frequencia_anterior': '',
        'frequencia_atual': '',
        'versao_anterior': '',
        'versao_atual': '',
        'ano_referencia_anterior': '',
        'ano_referencia_atual': '',
        'mes_referencia_anterior': '',
        'mes_referencia_atual': '',
    })

    def pick(row, key):
        for suffix in ('_atual', '_anterior'):
            value = row.get(f'{key}{suffix}')
            formatted = format_raw_value(value)
            if formatted:
                return formatted
        return ''

    def pick_anterior(row, key):
        """Pega o valor com sufixo _anterior"""
        value = row.get(f'{key}_anterior')
        return format_raw_value(value) if value else ''

    def pick_atual(row, key):
        """Pega o valor com sufixo _atual"""
        value = row.get(f'{key}_atual')
        return format_raw_value(value) if value else ''

    comparacao = []
    for _, row in merged.iterrows():
        valor_anterior = float(row['valor_anterior'] or 0)
        valor_atual = float(row['valor_atual'] or 0)
        diferenca = round(valor_atual - valor_anterior, 2)
        variacao_pct = None
        if valor_anterior != 0:
            variacao_pct = round((diferenca / valor_anterior) * 100, 2)

        status = determinar_status_comparacao(valor_anterior, valor_atual)
        status_variacao = calcular_status_variacao(valor_anterior, valor_atual)
        rubrica_codigo = format_raw_value(row.get('rubrica')) or format_raw_value(row.get('rubrica_anterior')) or format_raw_value(row.get('rubrica_atual')) or ''

        # Construir referências usando ano e mês dos arquivos
        ano_anterior = pick_anterior(row, 'ano_referencia')
        mes_anterior = pick_anterior(row, 'mes_referencia')
        referencia_anterior_calc = _construir_referencia(ano_anterior, mes_anterior)
        
        ano_atual = pick_atual(row, 'ano_referencia')
        mes_atual = pick_atual(row, 'mes_referencia')
        referencia_atual_calc = _construir_referencia(ano_atual, mes_atual)

        item = {
            'empresa': row['empresa'],
            'matricula': row['matricula'],
            'nome_servidor': pick(row, 'nome_servidor'),
            'rubrica': rubrica_codigo,
            'rubrica_codigo': rubrica_codigo,
            'dc_rubrica': pick(row, 'dc_rubrica'),
            'cpf': pick(row, 'cpf'),
            'referencia_anterior': referencia_anterior_calc,
            'referencia_atual': referencia_atual_calc,
            'situacao_funcional': pick(row, 'situacao_funcional'),
            'status_servidor': pick(row, 'status_servidor'),
            'descricao_status': pick(row, 'descricao_status'),
            'cargo': pick(row, 'cargo'),
            'descricao_cargo': pick(row, 'descricao_cargo'),
            'data_admissao': pick(row, 'data_admissao'),
            'data_ingresso_ref_salarial': pick(row, 'data_ingresso_ref_salarial'),
            'data_afastamento': pick(row, 'data_afastamento'),
            'motivo_afastamento': pick(row, 'motivo_afastamento'),
            'motivo_desligamento': pick(row, 'motivo_desligamento'),
            'carga_horaria': pick(row, 'carga_horaria'),
            'carga_horaria_anterior': format_raw_value(row['carga_horaria_anterior']),
            'carga_horaria_atual': format_raw_value(row['carga_horaria_atual']),
            'carga_horaria_secundaria': pick(row, 'carga_horaria_secundaria'),
            'carga_horaria_secundaria_anterior': format_raw_value(row['carga_horaria_secundaria_anterior']),
            'carga_horaria_secundaria_atual': format_raw_value(row['carga_horaria_secundaria_atual']),
            'ref_vertical': pick(row, 'ref_vertical'),
            'ref_horizontal': pick(row, 'ref_horizontal'),
            'prov_desc': pick(row, 'prov_desc'),
            'valor_mes_anterior': pick(row, 'valor_mes_anterior'),
            'valor_mes_atual': pick(row, 'valor_mes_atual'),
            'valor_vencimento': pick(row, 'valor_vencimento'),
            'valor_total_recebido': pick(row, 'valor_total_recebido'),
            'frequencia': pick(row, 'frequencia'),
            'versao_anterior': format_raw_value(row['versao_anterior']),
            'versao_atual': format_raw_value(row['versao_atual']),
            'valor_anterior': valor_anterior,
            'valor_atual': valor_atual,
            'diferenca': diferenca,
            'variacao_pct': variacao_pct,
            'status': status,
            'status_variacao': status_variacao,
        }
        comparacao.append(item)

    if not comparacao:
        return {'erro': 'Nenhuma comparação foi encontrada para as rubricas informadas.'}

    comparacao = sorted(comparacao, key=lambda x: (x['empresa'], x['matricula'], x.get('rubrica_codigo', '')))

    return {
        'comparacao': comparacao,
    }


def extract_year(s):
    """Extrai ano de 4 dígitos"""
    if not isinstance(s, str):
        s = str(s)
    match = re.search(r'\b(20\d{2}|19\d{2})\b', s)
    return int(match.group(1)) if match else None


def extract_number(s):
    """Extrai número inteiro"""
    if not isinstance(s, (int, float)):
        s = str(s)
        match = re.search(r'\d+', s.replace('.', '').replace(',', ''))
        return int(match.group(0)) if match else 0
    return int(s)


def normalize_reference(value):
    """Normaliza referências para comparação.

    Remove o sufixo A final em referências como S3A ou 25A, e também trunca
    sequências numéricas maiores que duas casas.
    """
    if value is None:
        return ''
    s = str(value).strip().upper()
    if s == '':
        return ''

    # Trata representações numéricas como 2.0 ou 3,0 corretamente
    s = s.replace(',', '.')
    num_match = re.fullmatch(r'(\d+)(?:\.0+)?', s)
    if num_match:
        digits = num_match.group(1)
        return digits[:2] if len(digits) > 2 else digits

    s = re.sub(r'[^A-Z0-9]', '', s)

    # Trata valores como S01 ou 001 como S1 e 1, removendo zeros à esquerda.
    match = re.fullmatch(r'([A-Z]+)(0*\d+)(A?)', s)
    if match:
        prefix, digits, suffix_a = match.groups()
        digits = digits.lstrip('0') or '0'
        normalized = prefix + digits
        if suffix_a and len(digits) <= 2:
            return normalized
        s = normalized + suffix_a

    # Remove o sufixo A final apenas em referências curtas como S3A ou 25A.
    if len(s) == 3 and s.endswith('A'):
        return s[:-1]

    def _truncate_digits(match):
        digits = match.group(0)
        return digits[:2]

    return re.sub(r'\d{3,}', _truncate_digits, s)


def norm_num(x):
    """Normaliza número para comparação"""
    try:
        import pandas as pd
        if pd.isna(x):
            return None
    except ImportError:
        pass
    try:
        x = float(str(x).replace(',', '.'))
        return round(x, 2)
    except:
        return None


def _prepare_formula_expression(expression):
    """Prepara expressão de fórmula em texto para avaliação segura."""
    if not expression:
        return None
    expr = str(expression).strip()
    expr = expr.replace('×', '*').replace('X', '*').replace('x', '*').replace('–', '-').replace('—', '-')
    expr = expr.replace('R$', '').replace('$', '')
    expr = expr.replace('(', ' ( ').replace(')', ' ) ')
    expr = re.sub(r'(?<![\d\.])(\d+),(\d+)', r'\1.\2', expr)
    expr = re.sub(r'(\d+(?:\.\d+)?)\s*%', r'(\1/100)', expr)
    expr = expr.replace('(', ' ( ').replace(')', ' ) ')

    replacements = {
        'valor vencimento': 'valor_vencimento',
        'valor_vencimento': 'valor_vencimento',
        'vencimento': 'valor_vencimento',
        'salario': 'valor_vencimento',
        'salário': 'valor_vencimento',
        'provento': 'valor_vencimento',
        'provento': 'valor_vencimento',
        'remuneracao': 'valor_vencimento',
        'remuneração': 'valor_vencimento',
        'valor recebido': 'valor_extrator',
        'valor_extrator': 'valor_extrator',
        'valor total recebido': 'valor_extrator',
        'valor absoluto': 'valor_extrator',
        'frequencia': 'frequencia',
        'frequência': 'frequencia',
        'porcentagem da frequencia': 'frequencia_percentual',
        'porcentagem da frequência': 'frequencia_percentual',
        'percentual da frequencia': 'frequencia_percentual',
        'percentual da frequência': 'frequencia_percentual',
        'frequencia percentual': 'frequencia_percentual',
        'frequência percentual': 'frequencia_percentual',
        'carga horaria': 'carga_horaria',
        'carga_horária': 'carga_horaria',
        'carga': 'carga_horaria',
        'ano referencia': 'ano',
        'ano_referencia': 'ano',
        'ano': 'ano',
        'ref_vertical': 'ref_vertical',
        'ref_horizontal': 'ref_horizontal',
    }
    for old, new in replacements.items():
        expr = re.sub(r'\b' + re.escape(old) + r'\b', new, expr, flags=re.IGNORECASE)

    # Remove duplicated whitespace
    expr = re.sub(r'\s+', ' ', expr).strip()
    return expr


def _extract_formula_candidate(expression):
    """Tenta recortar apenas a parte matemática de um texto descritivo."""
    if not expression:
        return None

    expr = _prepare_formula_expression(expression)
    if not expr:
        return None

    expr = expr.strip()

    token_pattern = re.compile(
        r'^(?:'
        r'[A-Za-z_][A-Za-z0-9_]*'
        r'|\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)?'
        r'|\('
        r'|\)'
        r'|\+'
        r'|\-'
        r'|\*'
        r'|/'
        r'|>=|<=|==|!='
        r')$'
    )

    tokens = expr.split()
    candidate_tokens = []
    for token in tokens:
        if token_pattern.match(token):
            candidate_tokens.append(token)
        else:
            break

    if candidate_tokens:
        return ' '.join(candidate_tokens)

    return expr


def safe_eval_formula(formula, context):
    """Avalia uma expressão de fórmula em um contexto seguro."""
    expr = _prepare_formula_expression(formula)
    if not expr:
        return None

    safe_context = {
        'abs': abs,
        'round': round,
        'min': min,
        'max': max,
        'pow': pow,
        'math': math,
    }
    safe_context.update({k: v for k, v in context.items() if v is not None})

    try:
        result = eval(expr, {'__builtins__': None}, safe_context)
    except Exception:
        return None

    if isinstance(result, bool):
        return float(result)
    if isinstance(result, (int, float)):
        return round(float(result), 2)
    return None


def _limpar_texto_rubrica(value):
    if value is None:
        return ''
    text = str(value).strip()
    if text.lower() == 'nan':
        return ''
    return text


def _normalizar_frequencia_percentual(frequencia):
    if frequencia is None:
        return None
    try:
        valor = float(frequencia)
    except Exception:
        return None

    if valor > 1:
        return round(valor / 100, 4)
    return round(valor, 4)


def _calcular_valor_esperado_rubrica_10020(valor_vencimento, frequencia, valor_extrator=None):
    if valor_vencimento is None:
        return None

    try:
        frequencia_num = float(frequencia) if frequencia is not None else 30
    except Exception:
        frequencia_num = 30

    if frequencia_num == 0:
        frequencia_num = 30

    if valor_extrator is not None:
        try:
            if float(valor_extrator) > 0 and frequencia_num == 0:
                frequencia_num = 30
        except Exception:
            pass

    valor_diario = float(valor_vencimento) / 30
    return round((valor_diario * frequencia_num) * 0.25, 2)


def _mapear_grau_instrucao_para_percentual(grau_instrucao, rubrica=None):
    texto = _limpar_texto_rubrica(grau_instrucao)
    if not texto:
        return None

    texto_norm = normalize(texto)
    texto_norm = re.sub(r'\s+', ' ', texto_norm).strip()

    if str(rubrica).strip() == '11033':
        mapeamento = {
            'ensino medio': 0.09,
            '2 graduacao': 0.09,
            '2a graduacao': 0.09,
            'segunda graduacao': 0.09,
            'graduacao': 0.13,
            'especializacao': 0.20,
            'mestrado': 0.30,
            'doutorado': 0.35,
        }
    elif str(rubrica).strip() == '11110':
        mapeamento = {
            '2 graduacao': 0.13,
            '2a graduacao': 0.13,
            'segunda graduacao': 0.13,
            'especializacao': 0.20,
            'mestrado': 0.30,
            'doutorado': 0.35,
        }
    elif str(rubrica).strip() == '11171':
        mapeamento = {
            'pos graduacao': 0.15,
            'graduacao': 0.15,
            'mestrado': 0.35,
            'doutorado': 0.40,
        }
    elif str(rubrica).strip() in {'11189', '11190'}:
        mapeamento = {
            '2 graduacao': 0.15,
            '2a graduacao': 0.15,
            'segunda graduacao': 0.15,
            'graduacao': 0.15,
            'especializacao': 0.25,
            'mestrado': 0.35,
            'doutorado': 0.40,
        }
    else:
        mapeamento = {
            'ensino medio completo': 0.10,
            'ensino superior completo': 0.15,
            'especializacao': 0.25,
            'mestrado': 0.35,
            'doutorado': 0.40,
        }

    for chave, percentual in mapeamento.items():
        if chave in texto_norm:
            return percentual

    return None


def _descrever_calculo_rubrica(rubrica_obj):
    if not rubrica_obj:
        return ''

    partes = []
    for label, attr in [
        ('Regra simplificada', 'regra_simplificada'),
        ('Critério de cálculo', 'criterio_calculo_rubrica'),
        ('Valor', 'valor'),
        ('Regra de cálculo', 'formula_calculo'),
        ('Base de cálculo', 'base_calculo'),
    ]:
        texto = _limpar_texto_rubrica(getattr(rubrica_obj, attr, ''))
        if texto:
            partes.append(f'{label}: {texto}')

    return ' | '.join(partes)


def _normalizar_codigo_rubrica(value):
    texto = _limpar_texto_rubrica(value)
    if not texto:
        return ''
    texto = texto.replace('.0', '')
    if texto.isdigit():
        if len(texto) > 5:
            texto = texto[-5:]
        return texto.zfill(5) if len(texto) < 5 else texto

    grupos = re.findall(r'\d+', texto)
    if grupos:
        codigo = max(grupos, key=len)
        if len(codigo) > 5:
            codigo = codigo[-5:]
        return codigo.zfill(5) if len(codigo) < 5 else codigo

    return texto


def _regras_suplementares_usuario():
    texto_legal_10024_10025 = (
        'Aplica-se o disposto nesta Lei, no que couber, aos servidores aposentados e aos benefi­ciários de pensão vinculados à carreira Policiamento e Fiscalização de Trânsito cujos proventos tenham paridade com os servidores ativos.'
    )
    return {
        '10020': SimpleNamespace(
            nome='10020',
            codigo='10020',
            descricao='VENCIMENTO proporcional à frequência',
            criterio_calculo_rubrica='vencimento x 25% x frequencia / 30',
            valor='vencimento x 25% x frequencia / 30',
            regra_simplificada='vencimento x 25% x frequencia / 30',
            formula_calculo='valor_vencimento * 0.25 * frequencia / 30',
            base_calculo='valor_vencimento',
            valor_padrao=None,
        ),
        '10024': SimpleNamespace(
            nome='10024',
            codigo='10024',
            descricao='REPRESENTAÇÃO proporcional à frequência',
            criterio_calculo_rubrica=f'((valor_vencimento / 30) * frequencia) * 0.25 {texto_legal_10024_10025}',
            valor='((valor_vencimento / 30) * frequencia) * 0.25',
            regra_simplificada='((valor_vencimento / 30) * frequencia) * 0.25',
            formula_calculo='((valor_vencimento / 30) * frequencia) * 0.25',
            base_calculo='valor_vencimento',
            valor_padrao=None,
        ),
        '10025': SimpleNamespace(
            nome='10025',
            codigo='10025',
            descricao='REPRESENTAÇÃO proporcional à frequência',
            criterio_calculo_rubrica=f'((valor_vencimento / 30) * frequencia) * 0.25 {texto_legal_10024_10025}',
            valor='((valor_vencimento / 30) * frequencia) * 0.25',
            regra_simplificada='((valor_vencimento / 30) * frequencia) * 0.25',
            formula_calculo='((valor_vencimento / 30) * frequencia) * 0.25',
            base_calculo='valor_vencimento',
            valor_padrao=None,
        ),
        '10059': SimpleNamespace(
            nome='10059',
            codigo='10059',
            descricao='VENCIMENTO',
            criterio_calculo_rubrica='vencimento x 25%',
            valor='vencimento x 25%',
            regra_simplificada='vencimento x 25%',
            formula_calculo='valor_vencimento * 25%',
            base_calculo='valor_vencimento',
            valor_padrao=None,
        ),
        '10130': SimpleNamespace(
            nome='10130',
            codigo='10130',
            descricao='VENCIMENTO',
            criterio_calculo_rubrica='vencimento x 10%',
            valor='vencimento x 10%',
            regra_simplificada='vencimento x 10%',
            formula_calculo='valor_vencimento * 10%',
            base_calculo='valor_vencimento',
            valor_padrao=None,
        ),
    }


@lru_cache(maxsize=1)
def _carregar_regras_simplificadas_automaticas():
    regras = {}
    caminho = Path(__file__).resolve().parent.parent / 'Arquivos' / 'regras_simplificadas_automaticas.xlsx'
    if not caminho.exists():
        debug_print(f"⚠️ Arquivo de regras não encontrado: {caminho}")
        return regras

    try:
        import openpyxl
    except Exception as exc:
        debug_print(f"⚠️ Não foi possível importar openpyxl para carregar regras: {exc}")
        return regras

    try:
        workbook = openpyxl.load_workbook(caminho, data_only=True)
        sheet = workbook.active
        headers = [normalize(col) if col is not None else '' for col in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        indices = {header: idx for idx, header in enumerate(headers)}

        idx_rubrica = indices.get('rubrica')
        idx_criterio = indices.get('criterio de calculo da rubrica')
        idx_regra = indices.get('regra_simplificada')
        idx_descricao = indices.get('descricao da rubrica sigrh')

        if idx_rubrica is None or idx_regra is None:
            debug_print('⚠️ Cabeçalhos esperados não encontrados no arquivo de regras simplificadas.')
            return regras

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rubrica_codigo = _normalizar_codigo_rubrica(row[idx_rubrica] if idx_rubrica < len(row) else '')
            regra_simplificada = _limpar_texto_rubrica(row[idx_regra] if idx_regra < len(row) else '')
            if not rubrica_codigo or not regra_simplificada:
                continue
            if regra_simplificada.casefold() == 'análise manual necessária':
                continue

            criterio = _limpar_texto_rubrica(row[idx_criterio] if idx_criterio is not None and idx_criterio < len(row) else '')
            descricao = _limpar_texto_rubrica(row[idx_descricao] if idx_descricao is not None and idx_descricao < len(row) else '')
            regras[rubrica_codigo] = SimpleNamespace(
                nome=rubrica_codigo,
                codigo=rubrica_codigo,
                descricao=descricao,
                criterio_calculo_rubrica=criterio or regra_simplificada,
                valor=regra_simplificada,
                regra_simplificada=regra_simplificada,
                formula_calculo=regra_simplificada,
                base_calculo='valor_vencimento',
                valor_padrao=None,
            )

        regras.update(_regras_suplementares_usuario())

        debug_print(f"✓ Regras automáticas carregadas: {len(regras)} rubricas")
        return regras
    except Exception as exc:
        debug_print(f"⚠️ Falha ao carregar regras simplificadas: {exc}")
        return regras


def _obter_rubrica_calculo(rubrica_codigo, rubrica_obj=None):
    if rubrica_obj is not None:
        return rubrica_obj

    codigo = _normalizar_codigo_rubrica(rubrica_codigo)
    if not codigo:
        return None

    regras_padrao = _carregar_regras_simplificadas_automaticas()
    regra = regras_padrao.get(codigo)
    if not regra:
        return None

    return regra


def _avaliar_calculo_rubrica(rubrica_obj, contexto):
    if not rubrica_obj:
        return None, None, None

    candidatos = [
        ('Regra simplificada', getattr(rubrica_obj, 'regra_simplificada', None)),
        ('Fórmula cadastrada', getattr(rubrica_obj, 'formula_calculo', None)),
        ('Critério de cálculo', getattr(rubrica_obj, 'criterio_calculo_rubrica', None)),
        ('Valor', getattr(rubrica_obj, 'valor', None)),
        ('Valor padrão', getattr(rubrica_obj, 'valor_padrao', None)),
    ]

    for origem, expressao in candidatos:
        if expressao in (None, ''):
            continue

        expressao_candidata = _extract_formula_candidate(expressao)
        for expressao_teste in [expressao_candidata, expressao]:
            valor_esperado = safe_eval_formula(expressao_teste, contexto)
            if valor_esperado is not None:
                return valor_esperado, origem, str(expressao_teste)

    return None, None, None


def parse_int_value(value):
    """Converte valor para inteiro preservando casos como '30,0' e '40.0'."""
    if value is None:
        return None
    try:
        s = str(value).strip().replace(',', '.')
        return int(float(s))
    except Exception:
        try:
            match = re.search(r'\d+', str(value))
            return int(match.group(0)) if match else None
        except Exception:
            return None


def parse_month_value(value):
    """Converte valores de mês para inteiro de 1 a 12."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return int(value)
        except Exception:
            return None
    s = str(value).strip().lower()
    if s == '':
        return None
    # Extrai dígitos simples como 1, 01 ou 12
    digits = re.search(r'\d+', s)
    if digits:
        try:
            month = int(digits.group(0))
            return month if 1 <= month <= 12 else None
        except Exception:
            pass
    # Normaliza nomes de mês
    meses = {
        'janeiro': 1, 'jan': 1,
        'fevereiro': 2, 'fev': 2,
        'março': 3, 'marco': 3, 'mar': 3,
        'abril': 4, 'abr': 4,
        'maio': 5,
        'junho': 6, 'jun': 6,
        'julho': 7, 'jul': 7,
        'agosto': 8, 'ago': 8,
        'setembro': 9, 'set': 9,
        'outubro': 10, 'out': 10,
        'novembro': 11, 'nov': 11,
        'dezembro': 12, 'dez': 12,
    }
    cleaned = re.sub(r'[^a-z]', '', s)
    return meses.get(cleaned)


def formatar_carga_horaria_recebida(carga_horaria_total, carga_horaria_primaria=None, carga_horaria_secundaria=None):
    """Formata a carga horária recebida, somando principal + secundária quando houver ambos os valores."""
    try:
        carga_total = int(carga_horaria_total) if carga_horaria_total is not None else None
    except Exception:
        carga_total = None

    carga_primaria = parse_int_value(carga_horaria_primaria) if carga_horaria_primaria is not None else None
    carga_secundaria = parse_int_value(carga_horaria_secundaria) if carga_horaria_secundaria is not None else None

    if carga_total is not None and carga_primaria is not None and carga_secundaria is not None:
        if carga_primaria + carga_secundaria == carga_total:
            return f"{carga_primaria}h + {carga_secundaria}h = {carga_total}h"

    if carga_total is not None:
        return f"{carga_total}h"
    if carga_primaria is not None:
        return f"{carga_primaria}h"
    return None


def find_alternate_carga_match(df_vencimento, col_refv_vertical, col_refv_horizontal,
                               col_ano_vencimento, col_carga_vencimento,
                               col_valor_vencimento, ref_v, ref_h, ano,
                               carga_horaria, valor_extrator):
    """Procura no VENCIMENTO um valor que corresponda a outra carga horária."""
    if valor_extrator is None:
        return None

    mask_alt = (
        (df_vencimento[col_refv_vertical].astype(str).apply(normalize_reference) == ref_v) &
        (df_vencimento[col_refv_horizontal].astype(str).apply(normalize_reference) == ref_h) &
        (df_vencimento[col_ano_vencimento].astype(str).str.replace(',', '') == str(ano))
    )

    df_alternativa = df_vencimento[mask_alt].copy()
    if df_alternativa.empty:
        return None

    for _, row_alt in df_alternativa.iterrows():
        carga_alt = parse_int_value(row_alt[col_carga_vencimento])
        if carga_alt is None or carga_alt == carga_horaria:
            continue

        valor_alt = norm_num(row_alt[col_valor_vencimento])
        if valor_alt is None:
            continue

        if abs(valor_alt - valor_extrator) <= 0.01:
            return {
                'carga': carga_alt,
                'valor': valor_alt,
            }

    return None


def parse_vigencia_date(value):
    """Converte DATA DE VIGÊNCIA para datetime, suportando formatos compactos como 1092014."""
    import pandas as pd

    if value is None:
        return pd.NaT
    s = str(value).strip()
    if s == '':
        return pd.NaT

    # Tenta converter diretamente com dayfirst
    try:
        parsed = pd.to_datetime(s, dayfirst=True, errors='coerce')
        if not pd.isna(parsed):
            return parsed
    except Exception:
        pass

    # Remove tudo que não for dígito e tenta parsear como ddmmaaaa
    digits = ''.join(c for c in s if c.isdigit())
    if len(digits) in (7, 8):
        digits = digits.zfill(8)
        try:
            return pd.to_datetime(digits, format='%d%m%Y', errors='coerce')
        except Exception:
            pass

    return pd.NaT


def read_excel_flexible(filepath):
    """
    Lê arquivo Excel com suporte a múltiplos formatos (.xls, .xlsx)
    Tenta diferentes engines automaticamente
    Detecta e tenta ler HTML se arquivo for HTML em vez de Excel
    """
    import pandas as pd
    import warnings
    
    filepath_str = str(filepath).lower()
    
    # Verifica se o arquivo é HTML (pode ter extensão .xls/.xlsx mas ser HTML)
    try:
        with open(filepath, 'rb') as f:
            first_bytes = f.read(100)
            if b'<html' in first_bytes.lower() or b'<?xml' in first_bytes[:50].lower():
                debug_print(f"  ℹ Arquivo detectado como HTML/XML, tentando ler tabelas...")
                try:
                    # Tenta ler a primeira tabela HTML
                    tables = pd.read_html(filepath)
                    if tables:
                        debug_print(f"  ✓ Sucesso ao ler tabela HTML (encontradas {len(tables)} tabelas)")
                        return tables[0]  # Retorna a primeira tabela
                except Exception as html_err:
                    debug_print(f"  ✗ Falha ao ler como HTML: {str(html_err)[:80]}")
                    raise Exception(
                        f"Arquivo {filepath} é HTML em vez de Excel real.\n"
                        f"Possível causa: arquivo baixado de página web que retornou HTML.\n"
                        f"Solução: Salve o arquivo diretamente do Excel ou exporte em formato .xlsx válido."
                    )
    except Exception as e:
        if "Arquivo" in str(e) and "HTML" in str(e):
            raise e  # Re-raise nossas mensagens de erro
        # Continua com tentativa normal de leitura
        pass
    
    # Tenta com engine apropriado baseado na extensão
    # IMPORTANTE: xlrd < 2.0 suporta .xls | openpyxl suporta .xlsx
    engines = []
    if filepath_str.endswith('.xlsx'):
        engines = ['openpyxl']  # openpyxl é obrigatório para .xlsx
    elif filepath_str.endswith('.xls'):
        engines = ['xlrd']  # xlrd < 2.0 é necessário para .xls
    else:
        # Se extensão desconhecida, tenta com base no arquivo
        engines = ['openpyxl', 'xlrd']
    
    last_error = None
    for engine in engines:
        try:
            debug_print(f"  Tentando ler com engine: {engine}")
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', UserWarning)
                df = pd.read_excel(filepath, engine=engine)
            debug_print(f"  ✓ Sucesso com engine: {engine}")
            return df
        except Exception as e:
            last_error = e
            debug_print(f"  ✗ Falha com {engine}: {str(e)[:80]}")
            continue
    
    # Se chegou aqui, nenhum engine funcionou
    raise Exception(f"Não foi possível ler o arquivo {filepath}. Formatos suportados: .xls, .xlsx. Erro: {str(last_error)}")


def processar_verificacao(file_vencimento, file_extrator, rubrica: str, ano: int, carga_horaria: int, rubrica_obj=None) -> Dict:
    """
    Processa dois arquivos Excel comparando conformidade de pagamentos.
    
    FLUXO:
    1. Filtrar EXTRATOR por: PROV/DESC = rubrica, ANO REFERENCIA = ano, CARGA HORARIA = carga_horaria
    2. Para cada pessoa no EXTRATOR filtrado:
       - Obter REF SALARIAL VERTICAL e REF SALARIAL HORIZONTAL
       - Procurar no VENCIMENTO por REFERENCIA DE VENCIMENTO VERTICAL e HORIZONTAL
       - Comparar VALOR DO VENCIMENTO com VALOR (EXTRATOR)
       - Quando rubrica = 10926, o VENCIMENTO não é necessário
    3. Retornar lista com pessoas, valores esperados, valores encontrados e status
    
    Args:
        file_vencimento: Arquivo VENCIMENTO (tabela de valores por referência)
        file_extrator: Arquivo EXTRATOR (folha de pagamento com pessoas)
        rubrica: PROV/DESC a filtrar
        ano: ANO REFERENCIA a filtrar
        carga_horaria: CARGA HORARIA a filtrar
        
    Returns:
        Dict com resultados: sucesso, resultados[], total, corretos, incorretos
    """
    try:
        import pandas as pd
    except ImportError as e:
        return {'erro': f'Pandas não foi configurado: {str(e)}. Reinstale com: pip install -r requirements.txt'}
    
    rubrica_str = str(rubrica).strip()
    rubrica_eh_10926 = rubrica_str == '10926'
    rubrica_eh_10014 = rubrica_str == '10014'
    rubrica_eh_11187 = rubrica_str == '11187'

    try:
        # Lê os arquivos com suporte a .xls e .xlsx
        if not rubrica_eh_10926 and not rubrica_eh_10014 and not rubrica_eh_11187:
            debug_print(f"\nLendo arquivo VENCIMENTO: {file_vencimento}")
            df_vencimento = _read_uploaded_excel(file_vencimento)
        else:
            df_vencimento = pd.DataFrame()

        debug_print(f"\nLendo arquivo EXTRATOR: {file_extrator}")
        df_extrator = _read_uploaded_excel(file_extrator)
    except Exception as e:
        return {'erro': f'Erro ao ler arquivos: {str(e)}. Certifique-se de que são arquivos válidos .xls ou .xlsx'}

    # Limpa headers
    df_vencimento = flatten_and_clean_columns(df_vencimento)
    df_extrator = flatten_and_clean_columns(df_extrator)

    if rubrica_obj is None and rubrica not in (None, '', 'todos', 'todas', 'all'):
        try:
            from .models import Rubrica

            rubrica_lookup = str(rubrica).strip()
            rubrica_codigo = _normalizar_codigo_rubrica(rubrica_lookup)

            rubrica_obj = Rubrica.objects.filter(codigo__iexact=rubrica_lookup).first()
            if rubrica_obj is None and rubrica_codigo:
                rubrica_obj = Rubrica.objects.filter(codigo__iexact=rubrica_codigo).first()
            if rubrica_obj is None:
                rubrica_obj = Rubrica.objects.filter(nome__iexact=rubrica_lookup).first()
            if rubrica_obj is None and rubrica_codigo:
                rubrica_obj = Rubrica.objects.filter(nome__iexact=rubrica_codigo).first()

            if rubrica_obj is not None:
                debug_print(f"✓ Rubrica carregada do banco: {rubrica_obj.codigo} - {rubrica_obj.nome}")
        except Exception as exc:
            debug_print(f"⚠️ Não foi possível carregar a rubrica do banco: {exc}")

    # Filtra o arquivo de vencimento quando há coluna de Filtro Vencimento
    col_filtro_vencimento = find_by_words(df_vencimento, ['filtro', 'vencimento']) or find_by_words(df_vencimento, ['filtro'])
    if col_filtro_vencimento and col_filtro_vencimento in df_vencimento.columns:
        debug_print(f"\n--- FILTRANDO VENCIMENTO POR '{col_filtro_vencimento}' ---")
        filtro_mask = df_vencimento[col_filtro_vencimento].astype(str).str.contains(
            r'Rubrica ligada ao vencimento|rubrica ligada ao vencimento|sim|true',
            case=False,
            na=False
        )
        df_filtrado = df_vencimento[filtro_mask]
        if not df_filtrado.empty:
            debug_print(f"  ✓ {len(df_filtrado)} registros marcados como 'Rubrica ligada ao vencimento?'")
            df_vencimento = df_filtrado
        else:
            debug_print(f"  ⚠️ Nenhum registro marcado como 'Rubrica ligada ao vencimento?' encontrado. Usando todo o arquivo de vencimento.")

    debug_print(f"\n{'='*80}")
    debug_print(f"=== ANÁLISE DETALHADA DOS ARQUIVOS ===")
    debug_print(f"{'='*80}")
    
    debug_print(f"\n📊 ARQUIVO VENCIMENTO:")
    debug_print(f"  Shape: {df_vencimento.shape}")
    debug_print(f"  Colunas: {df_vencimento.columns.tolist()}")
    debug_print(f"  Primeiras 2 linhas:")
    debug_print(df_vencimento.head(2).to_string())
    
    debug_print(f"\n📊 ARQUIVO EXTRATOR (ppgg.xlsx):")
    debug_print(f"  Shape: {df_extrator.shape}")
    debug_print(f"  Colunas: {df_extrator.columns.tolist()}")
    debug_print(f"  Primeiras 2 linhas:")
    debug_print(df_extrator.head(2).to_string())
    
    debug_print(f"\n{'='*80}")

    # ENCONTRA COLUNAS
    # EXTRATOR
    col_nome_servidor = _find_nome_servidor_column(df_extrator)
    col_prov = _find_rubrica_column(df_extrator, debug=True)
    col_rubrica = _select_rubrica_code_column(df_extrator)

    if not col_prov:
        debug_print("\n⚠️ RUBRICA NÃO ENCONTRADA - Tentando alternativas manuais...")
        for col in df_extrator.columns:
            if _is_amount_like_column_name(col):
                continue
            col_norm = normalize(col)
            if any(term in col_norm for term in ['rubrica', 'prov', 'desc', 'dc']):
                debug_print(f"  Usando alternativa: '{col}'")
                col_prov = col
                break
    
    col_ano_extrator = find_by_words(df_extrator, ['ano', 'referencia'])
    col_mes_extrator = find_by_words(df_extrator, ['mes', 'referencia']) or find_by_words(df_extrator, ['mes'])
    col_carga_extrator = _find_carga_horaria_column(df_extrator)
    col_carga_extrator_secundaria = find_by_words(df_extrator, ['carga', 'horaria', 'secundaria']) or find_by_words(df_extrator, ['carga', 'horária', 'secundária'])
    col_ref_vertical = find_by_words(df_extrator, ['ref', 'salarial', 'vertical'])
    col_ref_horizontal = find_by_words(df_extrator, ['ref', 'salarial', 'horizontal'])

    # Detecção robusta para referências funcionais: tenta múltiplas variações
    col_ref_funcional_vertical = (
        find_by_words(df_extrator, ['ref', 'funcional', 'vertical']) or
        find_by_words(df_extrator, ['ref', 'vertical', 'funcional']) or
        find_by_words(df_extrator, ['ref', 'funcional']) or
        find_by_words(df_extrator, ['funcional', 'vertical']) or
        next((col for col in df_extrator.columns if 'func' in normalize(col) and 'vert' in normalize(col)), None)
    )

    col_ref_funcional_horizontal = (
        find_by_words(df_extrator, ['ref', 'funcional', 'horizontal']) or
        find_by_words(df_extrator, ['ref', 'horizontal', 'funcional']) or
        find_by_words(df_extrator, ['ref', 'funcional']) or
        find_by_words(df_extrator, ['funcional', 'horizontal']) or
        next((col for col in df_extrator.columns if 'func' in normalize(col) and ('horiz' in normalize(col) or 'hor' in normalize(col))), None)
    )
    col_valor_extrator = next((col for col in ['VL_RUBRICA', 'VLRUBRICA', 'VALOR'] if col in df_extrator.columns), None)
    if not col_valor_extrator:
        col_valor_extrator = find_by_words(df_extrator, ['valor'])
    col_frequencia = find_by_words(df_extrator, ['frequencia'])  # Para rubrica 10502
    col_grau_instrucao = find_by_words(df_extrator, ['grau', 'instrucao'])
    col_empresa = (find_by_words(df_extrator, ['empresa']) or 
                   find_by_words(df_extrator, ['cod', 'empresa']) or 
                   find_by_words(df_extrator, ['codigo', 'empresa']) or
                   find_by_words(df_extrator, ['cod_empresa']) or
                   find_by_words(df_extrator, ['código']))
    col_dc_empresa = (next((col for col in [
                        'DC_EMPRESA', 'DESCRICAO_EMPRESA', 'DESCRICAO DA EMPRESA',
                        'DESC_EMPRESA', 'NOME DO ORGAO', 'NOME DO ÓRGÃO',
                        'NOME_ORGAO', 'NOME ÓRGÃO'
                    ] if col in df_extrator.columns), None) or
                     find_by_words(df_extrator, ['dc', 'empresa']) or
                     find_by_words(df_extrator, ['descricao', 'empresa']) or
                     find_by_words(df_extrator, ['nome', 'orgao']))
    col_rubrica = _select_rubrica_code_column(df_extrator)
    col_orgao = (find_by_words(df_extrator, ['orgao']) or
                 find_by_words(df_extrator, ['unidade']) or
                 find_by_words(df_extrator, ['lotacao']) or
                 find_by_words(df_extrator, ['lotação']))
    if not col_orgao:
        for col in df_extrator.columns:
            col_norm = normalize(col)
            if any(term in col_norm for term in ['orgao', 'org', 'órg', 'unidade', 'lotac']):
                col_orgao = col
                break
    col_cpf = find_by_words(df_extrator, ['cpf']) or find_by_words(df_extrator, ['documento']) or find_by_words(df_extrator, ['cpf/documento'])
    col_matricula = find_by_words(df_extrator, ['matricula', 'matrícula'])

    # Additional columns for detailed server information
    col_situacao_funcional = find_by_words(df_extrator, ['situacao', 'funcional']) or find_by_words(df_extrator, ['situação', 'funcional'])
    col_status = find_by_words(df_extrator, ['STATUS'])
    col_descricao_status = find_by_words(df_extrator, ['descricao', 'status']) or find_by_words(df_extrator, ['descrição', 'status'])
    col_cargo = find_by_words(df_extrator, ['cargo']) or find_by_words(df_extrator, ['funcao'])
    col_descricao_cargo = find_by_words(df_extrator, ['descricao', 'cargo']) or find_by_words(df_extrator, ['descrição', 'cargo'])
    col_data_admissao = find_by_words(df_extrator, ['data', 'admissao']) or find_by_words(df_extrator, ['data', 'admissão'])
    col_data_ingresso_ref_salarial = find_by_words(df_extrator, ['data', 'ingresso', 'referencia', 'salarial']) or find_by_words(df_extrator, ['data', 'ingresso', 'ref', 'salarial'])
    col_data_afastamento = find_by_words(df_extrator, ['data', 'afastamento'])
    col_motivo_afastamento = find_by_words(df_extrator, ['motivo', 'afastamento'])
    col_motivo_desligamento = find_by_words(df_extrator, ['motivo', 'desligamento']) or find_by_words(df_extrator, ['motivo', 'demissão'])
    col_carga_horaria = _find_carga_horaria_column(df_extrator)
    col_carga_horaria_secundaria = find_by_words(df_extrator, ['carga', 'horaria', 'secundaria']) or find_by_words(df_extrator, ['carga', 'horária', 'secundária'])
    col_prov_desc = next((col for col in ['RUBRICA', 'PROV/DESC', 'CD_PROV_DESC', 'ID_PROV_DESC'] if col in df_extrator.columns), None)
    if not col_prov_desc:
        col_prov_desc = find_by_words(df_extrator, ['rubrica']) or find_by_words(df_extrator, ['prov', 'desc'])
    if not col_prov_desc:
        col_prov_desc = next((col for col in df_extrator.columns if _is_five_digit_column(df_extrator[col])), None)
    col_valor_mes_anterior = next((col for col in ['VL_RUBRICA', 'VLRUBRICA'] if col in df_extrator.columns), None) or find_by_words(df_extrator, ['valor', 'mes', 'anterior']) or find_by_words(df_extrator, ['valor', 'mês', 'anterior'])
    col_valor_mes_atual = next((col for col in ['VL_RUBRICA', 'VLRUBRICA'] if col in df_extrator.columns), None) or find_by_words(df_extrator, ['valor', 'mes', 'atual']) or find_by_words(df_extrator, ['valor', 'mês', 'atual'])

    # VENCIMENTO
    col_refv_vertical = (
        next((col for col in ['REFER_SALARIAL_VERTICAL', 'REFERENCIA_DE_VENCIMENTO_VERTICAL', 'REFERENCIA VENCIMENTO VERTICAL'] if col in df_vencimento.columns), None)
        or find_by_words(df_vencimento, ['referencia', 'vencimento', 'vertical'])
        or find_by_words(df_vencimento, ['refer', 'salarial', 'vertical'])
    )
    col_refv_horizontal = (
        next((col for col in ['REFER_SALARIAL_HORIZONTAL', 'REFERENCIA_DE_VENCIMENTO_HORIZONTAL', 'REFERENCIA VENCIMENTO HORIZONTAL'] if col in df_vencimento.columns), None)
        or find_by_words(df_vencimento, ['referencia', 'vencimento', 'horizontal'])
        or find_by_words(df_vencimento, ['refer', 'salarial', 'horizontal'])
    )
    col_ano_vencimento = (
        next((col for col in ['ANO_REFER', 'ANO_REFERENCIA'] if col in df_vencimento.columns), None)
        or find_by_words(df_vencimento, ['ano', 'referencia'])
    )
    col_carga_vencimento = (
        next((col for col in ['CARGA_HORARIA'] if col in df_vencimento.columns), None)
        or find_by_words(df_vencimento, ['carga', 'horaria'])
    )
    col_valor_vencimento = (
        next((col for col in ['VL_VENCIMENTO', 'VALOR_VENCIMENTO', 'VL_VENC'] if col in df_vencimento.columns), None)
        or find_by_words(df_vencimento, ['valor', 'vencimento'])
        or find_by_words(df_vencimento, ['valor'])
    )
    col_data_vigencia = (
        next((col for col in ['DATA_VIGENCIA', 'DATA DE VIGENCIA', 'DATA_VIGENCIA_FIM'] if col in df_vencimento.columns), None)
        or find_by_words(df_vencimento, ['data', 'vigencia'])
    )

    debug_print(f"\n=== COLUNAS ENCONTRADAS ===")
    debug_print(f"EXTRATOR:")
    debug_print(f"  Nome Servidor: {col_nome_servidor}")
    debug_print(f"  PROV/DESC: {col_prov}")
    debug_print(f"  Ano: {col_ano_extrator}")
    debug_print(f"  Mês: {col_mes_extrator}")
    debug_print(f"  Carga: {col_carga_extrator}")
    debug_print(f"  Ref Vertical: {col_ref_vertical}")
    debug_print(f"  Ref Horizontal: {col_ref_horizontal}")
    debug_print(f"  Ref Funcional Vertical: {col_ref_funcional_vertical}")
    debug_print(f"  Ref Funcional Horizontal: {col_ref_funcional_horizontal}")
    debug_print(f"  Valor: {col_valor_extrator}")
    debug_print(f"  Frequência: {col_frequencia} (para rubrica 10502)")
    rubricas_grau_instrucao = (
        '10512', '10556', '10579', '10582', '10584', '10605', '10606',
        '10742', '10879', '11033', '11110', '11164', '11171', '11188', '11189', '11190', '11194',
    )
    debug_print(f"  Grau Instrução: {col_grau_instrucao} (para rubricas {', '.join(rubricas_grau_instrucao)})")
    debug_print(f"  Empresa: {col_empresa}")
    debug_print(f"  Descrição Empresa: {col_dc_empresa}")
    debug_print(f"  Orgão: {col_orgao}")
    debug_print(f"  CPF: {col_cpf}")
    debug_print(f"  Matrícula: {col_matricula}")
    debug_print(f"  Situação Funcional: {col_situacao_funcional}")
    debug_print(f"  Status: {col_status}")
    debug_print(f"  Descrição Status: {col_descricao_status}")
    debug_print(f"  Cargo: {col_cargo}")
    debug_print(f"  Descrição Cargo: {col_descricao_cargo}")
    debug_print(f"  Data Admissão: {col_data_admissao}")
    debug_print(f"  Data Ingresso Ref Salarial: {col_data_ingresso_ref_salarial}")
    debug_print(f"  Data Afastamento: {col_data_afastamento}")
    debug_print(f"  Motivo Afastamento: {col_motivo_afastamento}")
    debug_print(f"  Motivo Desligamento: {col_motivo_desligamento}")
    debug_print(f"  Carga Horária: {col_carga_horaria}")
    debug_print(f"  Carga Horária Secundária: {col_carga_horaria_secundaria}")
    debug_print(f"  Carga Horária Total: {col_carga_extrator_secundaria and (col_carga_extrator + ' + ' + col_carga_extrator_secundaria) or col_carga_extrator}")
    debug_print(f"  PROV/DESC: {col_prov_desc}")
    debug_print(f"  Valor Mês Anterior: {col_valor_mes_anterior}")
    debug_print(f"  Valor Mês Atual: {col_valor_mes_atual}")
    debug_print(f"VENCIMENTO:")
    debug_print(f"  Ref Vertical: {col_refv_vertical}")
    debug_print(f"  Ref Horizontal: {col_refv_horizontal}")
    debug_print(f"  Ano: {col_ano_vencimento}")
    debug_print(f"  Carga: {col_carga_vencimento}")
    debug_print(f"  Valor: {col_valor_vencimento}")
    debug_print(f"  DATA DE VIGENCIA: {col_data_vigencia}")

    # INFORMAÇÕES SOBRE OS ARQUIVOS (APÓS ENCONTRAR COLUNAS)
    print(f"\n=== INFORMAÇÕES SOBRE OS ARQUIVOS ===")
    print(f"VENCIMENTO: {df_vencimento.shape[0]} linhas, {df_vencimento.shape[1]} colunas")
    if col_ano_vencimento:
        anos_venc = sorted(pd.to_numeric(df_vencimento[col_ano_vencimento], errors='coerce').dropna().unique().astype(int).tolist())
        print(f"  Anos disponíveis: {anos_venc}")
    if col_carga_vencimento:
        cargas_venc = sorted(pd.to_numeric(df_vencimento[col_carga_vencimento], errors='coerce').dropna().unique().astype(int).tolist())
        print(f"  Cargas disponíveis: {cargas_venc}")
    
    print(f"\nEXTRATOR: {df_extrator.shape[0]} linhas, {df_extrator.shape[1]} colunas")
    if col_prov:
        print(f"  Rubricas (PROV/DESC) disponíveis: {df_extrator[col_prov].unique().tolist()[:20]}")
    if col_ano_extrator:
        anos_ext = sorted(pd.to_numeric(df_extrator[col_ano_extrator], errors='coerce').dropna().unique().astype(int).tolist())
        print(f"  Anos disponíveis: {anos_ext}")
    if col_carga_extrator:
        cargas_ext = sorted(pd.to_numeric(df_extrator[col_carga_extrator], errors='coerce').dropna().unique().astype(int).tolist())
        print(f"  Cargas disponíveis: {cargas_ext}")
    
    print(f"\n⚠️ ATENÇÃO: Você procura: rubrica={rubrica}, ano={ano}, carga={carga_horaria}\n")

    # AVISO se coluna EMPRESA não for encontrada
    if not col_empresa:
        print(f"\n⚠️ AVISO: Coluna EMPRESA não foi encontrada no EXTRATOR!")
        print(f"   Colunas disponíveis: {df_extrator.columns.tolist()}")
        print(f"   O sistema tentou procurar por: 'empresa', 'cod_empresa', 'código'")
        print(f"   Se você tem uma coluna com empresa/código, renomeie para um desses nomes.")
    else:
        print(f"\n✓ Coluna EMPRESA encontrada: {col_empresa}")
        print(f"  Exemplos de valores: {df_extrator[col_empresa].dropna().unique().tolist()[:5]}")

    # VALIDAÇÕES
    erros = []
    if not col_nome_servidor:
        erros.append("Coluna NOME DO SERVIDOR não encontrada")
    if not col_prov:
        erros.append("Coluna PROV/DESC não encontrada")
    if not col_ano_extrator:
        erros.append("Coluna ANO REFERENCIA (EXTRATOR) não encontrada")
    if not col_carga_extrator and not col_carga_extrator_secundaria:
        erros.append("Coluna CARGA HORARIA (EXTRATOR) ou CARGA HORARIA SECUNDARIA não encontrada")
    if rubrica_eh_10014 or rubrica_eh_11187:
        if not col_ref_funcional_vertical or not col_ref_funcional_horizontal:
            erros.append("Colunas REF FUNCIONAL VERTICAL/HORIZONTAL não encontradas")
    else:
        if not col_ref_vertical or not col_ref_horizontal:
            erros.append("Colunas REF SALARIAL VERTICAL/HORIZONTAL não encontradas")
    if not col_valor_extrator:
        erros.append("Coluna VALOR (EXTRATOR) não encontrada")
    if rubrica_eh_10926 or rubrica_eh_10014 or rubrica_eh_11187:
        # Para rubrica 10926, 10014 ou 11187, não precisamos do arquivo VENCIMENTO
        pass
    else:
        if str(rubrica).strip() in rubricas_grau_instrucao and not col_grau_instrucao:
            erros.append(f"Coluna DC_GRAU_INSTRUCAO não encontrada para as rubricas {', '.join(rubricas_grau_instrucao)}")
        if not col_refv_vertical or not col_refv_horizontal:
            erros.append("Colunas REFERENCIA DE VENCIMENTO VERTICAL/HORIZONTAL não encontradas")
        if not col_ano_vencimento:
            erros.append("Coluna ANO REFERENCIA (VENCIMENTO) não encontrada")
        if not col_carga_vencimento:
            erros.append("Coluna CARGA HORARIA (VENCIMENTO) não encontrada")
        if not col_valor_vencimento:
            erros.append("Coluna VALOR (VENCIMENTO) não encontrada")
    
    if erros:
        msgs = " | ".join(erros)
        debug_print(f"\n❌ ERROS NA DETECÇÃO DE COLUNAS:")
        debug_print(f"{msgs}")
        debug_print(f"\nColunas disponíveis no EXTRATOR:")
        debug_print(f"{df_extrator.columns.tolist()}")
        debug_print(f"\nColunas disponíveis no VENCIMENTO:")
        debug_print(f"{df_vencimento.columns.tolist()}")
        return {'erro': msgs}
    
    # AVISO sobre coluna de data de vigência
    if not col_data_vigencia:
        debug_print(f"\n⚠️ AVISO: Coluna DATA DE VIGÊNCIA não encontrada no VENCIMENTO.")
        debug_print(f"   Quando houver múltiplos registros, será usado o primeiro encontrado.")
        debug_print(f"   Para usar o valor mais recente, certifique-se de que há uma coluna 'DATA DE VIGENCIA' no arquivo.")

    # PASSO 1: FILTRAR EXTRATOR
    df_extrator_original = df_extrator.copy()
    debug_print(f"\n{'='*80}")
    debug_print(f"=== FILTRANDO EXTRATOR ===")
    debug_print(f"{'='*80}")
    debug_print(f"Antes de filtros: {len(df_extrator)} linhas")
    
    # Filtro por Rubrica
    debug_print(f"\n--- FILTRO RUBRICA ---")
    filtro_col = col_rubrica if col_rubrica and str(rubrica).strip().isdigit() else col_prov
    debug_print(f"Coluna usada para filtro de rubrica: {filtro_col}")
    debug_print(f"Todos valores únicos em {filtro_col}: {df_extrator[filtro_col].unique().tolist()}")
    debug_print(f"Procurando: '{rubrica}' (tipo: {type(rubrica).__name__})")
    
    # Converter para string, remover espaços e verificar
    df_extrator_prov = df_extrator[filtro_col].astype(str).str.strip()
    debug_print(f"Primeiros 20 valores após str().strip(): {df_extrator_prov.head(20).tolist()}")
    
    # Preparar rubrica para busca
    if rubrica is None:
        rubrica_str = ''
    else:
        rubrica_str = str(rubrica).strip()

    rubrica_eh_todas = rubrica_str.lower() in ('', 'todos', 'todas', 'all')
    debug_print(f"\nRubrica procurada: '{rubrica_str}'")

    if rubrica_eh_todas:
        debug_print("\n⚠️ Nenhuma rubrica informada ou indicação de todas as rubricas. A verificação será feita para todas as rubricas no arquivo de extrator.")
    else:
        # Testar diferentes tipos de match
        debug_print(f"\nTestando diferentes tipos de match:")
        mask_exato = df_extrator_prov == rubrica_str
        debug_print(f"  Igualdade exata (==): {mask_exato.sum()} matches")
        
        mask_contains = df_extrator_prov.str.contains(rubrica_str, case=False, na=False)
        debug_print(f"  Contains (case-insensitive): {mask_contains.sum()} matches")
        
        mask_contains_case = df_extrator_prov.str.contains(rubrica_str, case=True, na=False)
        debug_print(f"  Contains (case-sensitive): {mask_contains_case.sum()} matches")
        
        # Tenta primeiro igualdade exata (com strip)
        df_extrator_filtrado = df_extrator[mask_exato]
        debug_print(f"\nApós filtro rubrica={rubrica_str}: {len(df_extrator_filtrado)} linhas")
        
        # Se não encontrou, tenta contains (mais flexível)
        if len(df_extrator_filtrado) == 0:
            debug_print(f"\n⚠️ AVISO: Nenhum registro com rubrica exata '{rubrica_str}'")
            debug_print(f"   Tentando com contains (case-insensitive)...")
            df_extrator_filtrado = df_extrator[mask_contains]
            debug_print(f"   Encontrados: {len(df_extrator_filtrado)} linhas")
        
        # Se ainda não encontrou, tenta contains case-sensitive
        if len(df_extrator_filtrado) == 0:
            debug_print(f"\n   Tentando com contains (case-sensitive)...")
            df_extrator_filtrado = df_extrator[mask_contains_case]
            debug_print(f"   Encontrados: {len(df_extrator_filtrado)} linhas")
        
        if len(df_extrator_filtrado) == 0:
            debug_print(f"\n❌ Nenhum registro encontrado com rubrica='{rubrica_str}'")
            rubricas_disponiveis = df_extrator_prov.unique().tolist()[:50]
            debug_print(f"   Rubricas disponíveis: {rubricas_disponiveis}")
            return {'erro': f'Rubrica "{rubrica_str}" não encontrada. Disponíveis: {rubricas_disponiveis}'}

        df_extrator = df_extrator_filtrado
    
    # Filtro por Ano
    debug_print(f"\n--- FILTRO ANO ---")
    debug_print(f"Coluna usada: {col_ano_extrator}")
    debug_print(f"Valores crus em {col_ano_extrator}: {df_extrator[col_ano_extrator].head(20).tolist()}")
    debug_print(f"Procurando: {ano} (tipo: {type(ano).__name__})")
    
    try:
        df_extrator_ano = df_extrator[col_ano_extrator].astype(str).str.strip().str.replace(',', '').astype(float).astype(int)
        debug_print(f"Valores após conversão para int: {df_extrator_ano.unique().tolist()}")
    except Exception as e:
        debug_print(f"Erro na conversão: {e}, tentando extract...")
        df_extrator_ano = df_extrator[col_ano_extrator].astype(str).str.extract(r'(\d+)')[0].astype(int)
        debug_print(f"Valores após extract: {df_extrator_ano.unique().tolist()}")
    
    debug_print(f"\nComparando:")
    debug_print(f"  Procura-se: {ano} (tipo: {type(ano).__name__})")
    debug_print(f"  Disponíveis: {sorted(df_extrator_ano.unique().tolist())}")
    
    mask_ano = df_extrator_ano == ano
    debug_print(f"Matches encontrados: {mask_ano.sum()} linhas")
    df_extrator = df_extrator[mask_ano]
    debug_print(f"Após filtro ano={ano}: {len(df_extrator)} linhas")
    
    if len(df_extrator) == 0:
        debug_print(f"❌ Nenhum registro encontrado com ano={ano}")
        anos_disponiveis = sorted(df_extrator_ano.unique().astype(int).tolist())
        debug_print(f"   Anos disponíveis: {anos_disponiveis}")
        return {'erro': f'Ano {ano} não encontrado em EXTRATOR. Anos disponíveis: {anos_disponiveis}. Use um dos anos disponíveis!'}
    
    # Tenta extrair Mês da coluna do EXTRATOR, se disponível
    mes_referencia = None
    if col_mes_extrator and col_mes_extrator in df_extrator.columns:
        mes_values = df_extrator[col_mes_extrator].dropna().apply(parse_month_value)
        mes_values = mes_values.dropna().unique().tolist()
        if len(mes_values) == 1:
            mes_referencia = int(mes_values[0])
            debug_print(f"  Mês de referência extraído do EXTRATOR: {mes_referencia}")
        elif len(mes_values) > 1:
            mes_referencia = int(mes_values[0])
            debug_print(f"  Múltiplos meses encontrados em {col_mes_extrator}: {mes_values}. Usando o primeiro: {mes_referencia}")
        else:
            debug_print(f"  Coluna {col_mes_extrator} encontrada, mas não foi possível extrair mês válido.")

    # Filtro por Carga Horária
    debug_print(f"\n--- FILTRO CARGA HORÁRIA ---")
    debug_print(f"Coluna usada: {col_carga_extrator or col_carga_extrator_secundaria}")
    if col_carga_extrator and col_carga_extrator in df_extrator.columns:
        debug_print(f"Todos valores únicos em {col_carga_extrator}: {df_extrator[col_carga_extrator].unique().tolist()}")
    if col_carga_extrator_secundaria and col_carga_extrator_secundaria in df_extrator.columns:
        debug_print(f"Todos valores únicos em {col_carga_extrator_secundaria}: {df_extrator[col_carga_extrator_secundaria].unique().tolist()}")
    debug_print(f"Procurando: {carga_horaria} (tipo: {type(carga_horaria).__name__})")
    
    def parse_carga_int(value):
        parsed = parse_int_value(value)
        return parsed if parsed is not None else 0

    carga_primaria = df_extrator[col_carga_extrator].apply(parse_carga_int) if col_carga_extrator and col_carga_extrator in df_extrator.columns else 0
    carga_secundaria = df_extrator[col_carga_extrator_secundaria].apply(parse_carga_int) if col_carga_extrator_secundaria and col_carga_extrator_secundaria in df_extrator.columns else 0

    df_extrator['carga_horaria_total'] = carga_primaria + carga_secundaria

    debug_print(f"  Carga Horária Total usada para filtro: {carga_horaria}")
    debug_print(f"  Valores combinados de carga horária disponíveis: {sorted(df_extrator['carga_horaria_total'].dropna().unique().tolist())}")

    # A regra de agrupamento considera qualquer pessoa cuja carga total corresponda
    # ao grupo de referência solicitado. Ex.: 30 + 10 entra em 40h, 20 + 10 entra em 30h.
    mask_carga = df_extrator['carga_horaria_total'] == carga_horaria
    debug_print(f"Matches encontrados com carga total exata: {mask_carga.sum()} linhas")

    # Quando a carga horária principal tiver secundária, o agrupamento deve considerar
    # a soma da carga principal + secundária como pertencente ao grupo alvo.
    if not mask_carga.any() and col_carga_extrator_secundaria and col_carga_extrator_secundaria in df_extrator.columns:
        df_extrator['carga_horaria_total'] = carga_primaria + carga_secundaria
        # Mantém o comportamento anterior quando não houver correspondência exata, mas
        # recorre ao total somado para incluir o registro no bucket correto.
        mask_carga = df_extrator['carga_horaria_total'].isin([carga_horaria])

    debug_print(f"Matches finais encontrados: {mask_carga.sum()} linhas")
    df_extrator = df_extrator[mask_carga]
    debug_print(f"Após filtro carga={carga_horaria}: {len(df_extrator)} linhas")

    if df_extrator.empty:
        debug_print(f"❌ Nenhum registro encontrado com carga={carga_horaria}")
        cargas_disponiveis = sorted(pd.to_numeric(
            read_excel_flexible(file_extrator)[col_carga_extrator], 
            errors='coerce'
        ).dropna().unique().astype(int).tolist())
        debug_print(f"   Cargas disponíveis: {cargas_disponiveis}")
        return {'erro': f'Carga horária {carga_horaria} não encontrada em EXTRATOR. Cargas disponíveis: {cargas_disponiveis}'}
    
    debug_print(f"\n{'='*80}")
    debug_print(f"✅ FILTROS APLICADOS COM SUCESSO!")
    debug_print(f"  Rubrica: {rubrica}")
    debug_print(f"  Ano: {ano}")
    debug_print(f"  Carga: {carga_horaria}")
    debug_print(f"  Total de registros para processar: {len(df_extrator)}")
    debug_print(f"{'='*80}\n")

    # PASSO 2: COMPARAR PESSOAS
    debug_print(f"\n=== COMPARANDO VALORES ===")
    resultados = []
    corretos = 0
    incorretos = 0
    verificar = 0

    for idx, row_extrator in df_extrator.iterrows():
        nome_servidor = row_extrator[col_nome_servidor]
        if rubrica_eh_10014 or rubrica_eh_11187:
            ref_v_raw = format_raw_value(row_extrator[col_ref_funcional_vertical]) if col_ref_funcional_vertical else ''
            ref_h_raw = format_raw_value(row_extrator[col_ref_funcional_horizontal]) if col_ref_funcional_horizontal else ''
            funcao_codigo = _obter_codigo_funcional(row_extrator, col_ref_funcional_vertical, col_ref_funcional_horizontal)
            ref_v = normalize_reference(ref_v_raw)
            ref_h = normalize_reference(ref_h_raw)
        else:
            ref_v_raw = row_extrator[col_ref_vertical]
            ref_h_raw = row_extrator[col_ref_horizontal]
            ref_v = normalize_reference(ref_v_raw)
            ref_h = normalize_reference(ref_h_raw)
        valor_extrator = norm_num(row_extrator[col_valor_extrator])
        frequencia = norm_num(row_extrator[col_frequencia]) if col_frequencia else None
        frequencia_percentual = _normalizar_frequencia_percentual(frequencia)
        carga_recebida_texto = formatar_carga_horaria_recebida(
            row_extrator['carga_horaria_total'] if 'carga_horaria_total' in row_extrator else None,
            row_extrator[col_carga_horaria] if col_carga_horaria and col_carga_horaria in df_extrator.columns else None,
            row_extrator[col_carga_horaria_secundaria] if col_carga_horaria_secundaria and col_carga_horaria_secundaria in df_extrator.columns else None,
        )
        grau_instrucao = format_raw_value(row_extrator[col_grau_instrucao]) if col_grau_instrucao else ''
        grau_instrucao_percentual = _mapear_grau_instrucao_para_percentual(grau_instrucao, rubrica)
        rubrica_codigo_linha = _limpar_texto_rubrica(format_raw_value(row_extrator[col_prov]) if col_prov and col_prov in df_extrator.columns else rubrica)

        rubrica_calculo_obj = _obter_rubrica_calculo(rubrica_codigo_linha, rubrica_obj=rubrica_obj)
        rubrica_descricao_calculo = _descrever_calculo_rubrica(rubrica_calculo_obj)
        is_aposentado = _is_servidor_aposentado(
            row_extrator,
            col_situacao_funcional=col_situacao_funcional,
            col_status=col_status,
            col_descricao_status=col_descricao_status,
        )
        
        # Extrai EMPRESA preservando zeros à esquerda
        if col_empresa:
            empresa_raw = str(row_extrator[col_empresa]).strip()
            # Remove ".0" se for número, depois preserva zeros à esquerda
            empresa = empresa_raw.replace('.0', '') if empresa_raw.endswith('.0') else empresa_raw
            # Garante zeros à esquerda (mínimo 3 dígitos para códigos como 007)
            if empresa.isdigit():
                empresa = empresa.zfill(3)
        else:
            empresa = ''

        if col_dc_empresa and col_dc_empresa in df_extrator.columns:
            dc_empresa = format_raw_value(row_extrator[col_dc_empresa])
        else:
            dc_empresa = ''
        
        # Extrai CPF preservando formato completo
        if col_cpf and col_cpf in df_extrator.columns:
            cpf = str(row_extrator[col_cpf]).strip()
            # Remove ".0" se for número
            cpf = cpf.replace('.0', '') if cpf.endswith('.0') else cpf
            cpf = cpf if cpf and cpf.lower() != 'nan' else ''
        else:
            cpf = ''
        
        matricula = str(row_extrator[col_matricula]).strip() if col_matricula else ''

        # Extrai órgão, se disponível
        if col_orgao and col_orgao in df_extrator.columns:
            orgao_raw = str(row_extrator[col_orgao]).strip()
            orgao = orgao_raw if orgao_raw and orgao_raw.lower() != 'nan' else ''
        else:
            orgao = ''

        valor_esperado = None

        if rubrica_eh_10926 or rubrica_eh_10014 or rubrica_eh_11187:
            registros_encontrados = pd.DataFrame()
            alternate_carga_match = None
        else:
            # Busca o valor esperado no VENCIMENTO
            # Precisa que REFERENCIA DE VENCIMENTO VERTICAL = ref_v
            #                  REFERENCIA DE VENCIMENTO HORIZONTAL = ref_h
            #                  ANO REFERENCIA = ano
            #                  CARGA HORARIA = carga_horaria
            
            mascara = (
                (df_vencimento[col_refv_vertical].astype(str).apply(normalize_reference) == ref_v) &
                (df_vencimento[col_refv_horizontal].astype(str).apply(normalize_reference) == ref_h) &
                (df_vencimento[col_ano_vencimento].astype(str).str.replace(',', '') == str(ano)) &
                (df_vencimento[col_carga_vencimento].apply(parse_int_value) == carga_horaria)
            )
            
            registros_encontrados = df_vencimento[mascara]
            
            # Se encontrou múltiplos registros e há coluna de data de vigência,
            # ordena por data de vigência descendente para pegar o mais recente
            if not registros_encontrados.empty and col_data_vigencia and col_data_vigencia in registros_encontrados.columns:
                try:
                    # Tenta converter para datetime e ordenar
                    registros_encontrados = registros_encontrados.copy()
                    registros_encontrados['data_vigencia_parsed'] = registros_encontrados[col_data_vigencia].apply(parse_vigencia_date)
                    registros_encontrados = registros_encontrados.sort_values(
                        'data_vigencia_parsed', 
                        ascending=False, 
                        na_position='last'
                    )
                    debug_print(f"  [DATA VIGÊNCIA] Ordenando {len(registros_encontrados)} registros por data de vigência")
                    debug_print(f"    Data mais recente: {registros_encontrados.iloc[0]['data_vigencia_parsed']}")
                except Exception as e:
                    debug_print(f"  [AVISO] Erro ao ordenar por data de vigência: {e}. Usando primeiro registro encontrado.")
            
            alternate_carga_match = find_alternate_carga_match(
                df_vencimento,
                col_refv_vertical,
                col_refv_horizontal,
                col_ano_vencimento,
                col_carga_vencimento,
                col_valor_vencimento,
                ref_v,
                ref_h,
                ano,
                carga_horaria,
                valor_extrator,
            )

        if registros_encontrados.empty and not rubrica_eh_10926 and not rubrica_eh_10014 and not rubrica_eh_11187:
            valor_vencimento = None
            if alternate_carga_match:
                status = 'INCORRETO'
                diferenca_absoluta = None
                diferenca_percentual = None
                justificativa = (
                    f'Carga horária divergente: recebeu como '
                    f'{carga_recebida_texto or f"{alternate_carga_match["carga"]}h"} '
                    f'(R$ {valor_extrator}) em vez de {carga_horaria}h.'
                )
                incorretos += 1
            else:
                status = 'INCORRETO'
                valor_vencimento = None
                diferenca_absoluta = None
                diferenca_percentual = None
                justificativa = (
                    f'Não foi encontrado vencimento para a combinação'
                    f'empresa={empresa}, órgão={orgao}, carga horária={carga_horaria},'
                    f'ref_vertical={ref_v}, ref_horizontal={ref_h}.'
                )
                incorretos += 1
        elif rubrica_eh_10926:
            valor_vencimento = None
            justificativa = None
            if frequencia is not None:
                frequencia_usada = min(frequencia, 22)
                valor_esperado = round((640 / 22) * frequencia_usada, 2)
                diferenca_absoluta = abs(valor_extrator - valor_esperado)
                diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                diferenca_percentual = round(diferenca_percentual, 2)
                diferenca_absoluta = round(diferenca_absoluta, 2)

                debug_print(f"\n[RUBRICA 10926] {nome_servidor}:")
                debug_print(f"  Frequência original: {frequencia}")
                debug_print(f"  Frequência usada: {frequencia_usada}")
                debug_print(f"  Valor Esperado: (640 / 22) × {frequencia_usada} = {valor_esperado}")
                debug_print(f"  Valor Extrator: {valor_extrator}")
                debug_print(f"  Diferença: {diferenca_absoluta} ({diferenca_percentual}%)")

                if valor_extrator > 640:
                    status = 'INCORRETO'
                    justificativa = f'Rubrica 10926: valor recebido maior que R$ 640 (R$ {valor_extrator}) é incorreto.'
                    incorretos += 1
                elif valor_extrator == valor_esperado or (diferenca_absoluta is not None and diferenca_absoluta <= 0.01):
                    status = 'CORRETO'
                    justificativa = (
                        f'Rubrica 10926: R$ {valor_extrator} está correto para frequência {frequencia} '
                        f'(usando {frequencia_usada} quando frequencia > 22) = R$ {valor_esperado} ✓'
                    )
                    corretos += 1
                elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                    status = 'VERIFICAR (Valores muito próximos)'
                    justificativa = f'Rubrica 10926: esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%)'
                    verificar += 1
                else:
                    status = 'INCORRETO'
                    justificativa = f'Rubrica 10926: esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator}.'
                    incorretos += 1
            else:
                status = 'INCORRETO'
                justificativa = 'Rubrica 10926: frequência não encontrada'
                incorretos += 1
                diferenca_absoluta = None
                diferenca_percentual = None
        elif rubrica_eh_10014:
            valor_vencimento = None
            justificativa = None
            funcao_codigo = _obter_codigo_funcional(row_extrator, col_ref_funcional_vertical, col_ref_funcional_horizontal)
            valor_esperado = _calcular_valor_rubrica_10014(funcao_codigo, frequencia)

            if valor_esperado is None:
                status = 'INCORRETO'
                justificativa = f'Rubrica 10014: função funcional "{funcao_codigo}" não encontrada no arquivo de legislação.'
                incorretos += 1
                diferenca_absoluta = None
                diferenca_percentual = None
            else:
                diferenca_absoluta = abs(valor_extrator - valor_esperado)
                diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                diferenca_percentual = round(diferenca_percentual, 2)
                diferenca_absoluta = round(diferenca_absoluta, 2)

                debug_print(f"\n[RUBRICA 10014] {nome_servidor}:")
                debug_print(f"  Função funcional: {funcao_codigo}")
                debug_print(f"  Valor esperado: R$ {valor_esperado}")
                debug_print(f"  Valor Extrator: R$ {valor_extrator}")
                debug_print(f"  Diferença: R$ {diferenca_absoluta} ({diferenca_percentual}%)")

                if valor_extrator == valor_esperado or diferenca_absoluta <= 0.01:
                    status = 'CORRETO'
                    justificativa = (
                        f'Rubrica 10014: função "{funcao_codigo}" utiliza o valor de remuneração do JSON '
                        f'({valor_esperado} = {valor_esperado * 0.80:.2f} / 0.80) ✓'
                    )
                    corretos += 1
                elif diferenca_percentual < 0.5:
                    status = 'VERIFICAR (Valores muito próximos)'
                    justificativa = f'Rubrica 10014: esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%)'
                    verificar += 1
                else:
                    status = 'INCORRETO'
                    justificativa = f'Rubrica 10014: esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator}.'
                    incorretos += 1
        elif rubrica_eh_11187:
            valor_vencimento = None
            justificativa = None
            funcao_codigo = _obter_codigo_funcional(row_extrator, col_ref_funcional_vertical, col_ref_funcional_horizontal)
            valor_esperado_10014 = _calcular_valor_rubrica_10014(funcao_codigo, frequencia)

            if rubrica_eh_11187 and valor_esperado_10014 is not None:
                valor_base_10014 = None
                chave_nome = normalize(_limpar_texto_rubrica(nome_servidor))
                chave_cpf = normalize(_limpar_texto_rubrica(cpf))
                chave_matricula = normalize(_limpar_texto_rubrica(matricula))
                chave_empresa = normalize(_limpar_texto_rubrica(empresa))
                chave_orgao = normalize(_limpar_texto_rubrica(orgao))
                chave_ano = normalize(_limpar_texto_rubrica(ano))
                chave_carga = normalize(_limpar_texto_rubrica(carga_horaria))

                if not df_extrator_original.empty and col_prov in df_extrator_original.columns:
                    linhas_10014 = df_extrator_original[
                        df_extrator_original[col_prov].astype(str).str.strip().str.contains(r'^10014$', case=False, na=False)
                    ].copy()
                    if not linhas_10014.empty:
                        for _, linha_10014 in linhas_10014.iterrows():
                            nome_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_nome_servidor]) if col_nome_servidor in linha_10014.index else ''))
                            cpf_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_cpf]) if col_cpf in linha_10014.index else ''))
                            matricula_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_matricula]) if col_matricula in linha_10014.index else ''))
                            empresa_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_empresa]) if col_empresa in linha_10014.index else ''))
                            orgao_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_orgao]) if col_orgao in linha_10014.index else ''))
                            ano_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_ano_extrator]) if col_ano_extrator in linha_10014.index else ''))
                            carga_10014 = normalize(_limpar_texto_rubrica(format_raw_value(linha_10014[col_carga_extrator]) if col_carga_extrator in linha_10014.index else ''))

                            same_name = chave_nome and nome_10014 and chave_nome == nome_10014
                            same_cpf = chave_cpf and cpf_10014 and chave_cpf == cpf_10014
                            same_matricula = chave_matricula and matricula_10014 and chave_matricula == matricula_10014
                            same_empresa = chave_empresa and empresa_10014 and chave_empresa == empresa_10014
                            same_orgao = chave_orgao and orgao_10014 and chave_orgao == orgao_10014
                            same_ano = chave_ano and ano_10014 and chave_ano == ano_10014
                            same_carga = chave_carga and carga_10014 and chave_carga == carga_10014

                            if ((same_name or same_cpf or same_matricula) and same_ano and same_carga and (same_empresa or same_orgao or not any([same_empresa, same_orgao]))):
                                valor_base_10014 = norm_num(linha_10014[col_valor_extrator])
                                break

                if valor_base_10014 is not None:
                    valor_esperado_10014 = valor_base_10014

            if valor_esperado_10014 is None:
                status = 'INCORRETO'
                justificativa = f'Rubrica 11187: função funcional "{funcao_codigo}" não encontrada no arquivo de legislação para derivar o valor da rubrica 10014.'
                incorretos += 1
                diferenca_absoluta = None
                diferenca_percentual = None
            else:
                valor_esperado = round(valor_esperado_10014 / 3, 2)
                diferenca_absoluta = abs(valor_extrator - valor_esperado)
                diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                diferenca_percentual = round(diferenca_percentual, 2)
                diferenca_absoluta = round(diferenca_absoluta, 2)

                debug_print(f"\n[RUBRICA 11187] {nome_servidor}:")
                debug_print(f"  Função funcional: {funcao_codigo}")
                debug_print(f"  Valor esperado da rubrica 10014: R$ {valor_esperado_10014}")
                debug_print(f"  Valor esperado da rubrica 11187: R$ {valor_esperado} (10014 / 3)")
                debug_print(f"  Valor Extrator: R$ {valor_extrator}")
                debug_print(f"  Diferença: R$ {diferenca_absoluta} ({diferenca_percentual}%)")

                if valor_extrator == valor_esperado or diferenca_absoluta <= 0.01:
                    status = 'CORRETO'
                    justificativa = (
                        f'Rubrica 11187: validação realizada apenas com o arquivo EXTRATOR. '
                        f'A rubrica 10014 foi considerada correta pela regra funcional e o valor da rubrica 11187 '
                        f'foi comparado com o valor esperado calculado como R$ {valor_esperado} '
                        f'(R$ {valor_esperado_10014} ÷ 3). '
                        f'Valor recebido: R$ {valor_extrator}. ✓'
                    )
                    corretos += 1
                elif diferenca_percentual < 0.5:
                    status = 'VERIFICAR (Valores muito próximos)'
                    justificativa = (
                        f'Rubrica 11187: validação realizada apenas com o arquivo EXTRATOR. '
                        f'Esperado R$ {valor_esperado} com base na rubrica 10014 (R$ {valor_esperado_10014} ÷ 3), '
                        f'mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%).'
                    )
                    verificar += 1
                else:
                    status = 'INCORRETO'
                    justificativa = (
                        f'Rubrica 11187: validação realizada apenas com o arquivo EXTRATOR. '
                        f'Esperado R$ {valor_esperado} com base na rubrica 10014 (R$ {valor_esperado_10014} ÷ 3), '
                        f'mas recebeu R$ {valor_extrator}.'
                    )
                    incorretos += 1
        else:
            valor_vencimento = norm_num(registros_encontrados.iloc[0][col_valor_vencimento])
            
            # Calcula diferença
            if valor_vencimento is not None and valor_extrator is not None:
                justificativa = None
                valor_esperado = None

                # Regra de baixa frequência para qualquer rubrica
                if frequencia is not None and frequencia <= 1 and valor_extrator != 0:
                    valor_esperado_alt_30 = round(valor_extrator * 30, 2)
                    valor_esperado_alt_31 = round(valor_extrator * 31, 2)
                    ratio = valor_vencimento / valor_extrator if valor_extrator != 0 else None
                    if ratio is not None and abs(ratio - 30) <= 0.1:
                        status = 'CORRETO'
                        justificativa = f'Baixa frequência: R$ {valor_extrator} × 30 ≈ R$ {valor_esperado_alt_30} ✓'
                        corretos += 1
                    elif ratio is not None and abs(ratio - 31) <= 0.1:
                        status = 'CORRETO'
                        justificativa = f'Baixa frequência: R$ {valor_extrator} × 31 ≈ R$ {valor_esperado_alt_31} ✓'
                        corretos += 1

                if justificativa is None:
                    if rubrica_codigo_linha == '10008' and is_aposentado:
                        valor_esperado = valor_vencimento
                        origem_calculo = 'Tabela de Vencimento'
                        expressao_calculo = 'valor_vencimento'
                    elif rubrica_calculo_obj is not None and rubrica_codigo_linha not in (*rubricas_grau_instrucao, '10502', '10020', '10024', '10025'):
                        contexto_calculo = {
                            'valor_vencimento': valor_vencimento,
                            'valor_extrator': valor_extrator,
                            'frequencia': frequencia,
                            'frequencia_percentual': frequencia_percentual,
                            'carga_horaria': carga_horaria,
                            'ano': ano,
                            'ref_vertical': ref_v,
                            'ref_horizontal': ref_h,
                            'empresa': empresa,
                            'cpf': cpf,
                            'matricula': matricula,
                            'valor_padrao': float(rubrica_calculo_obj.valor_padrao) if getattr(rubrica_calculo_obj, 'valor_padrao', None) is not None else None,
                        }
                        valor_esperado, origem_calculo, expressao_calculo = _avaliar_calculo_rubrica(
                            rubrica_calculo_obj,
                            contexto_calculo,
                        )
                        if valor_esperado is not None:
                            diferenca_absoluta = abs(valor_extrator - valor_esperado)
                            diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                            diferenca_percentual = round(diferenca_percentual, 2)
                            diferenca_absoluta = round(diferenca_absoluta, 2)

                            rubrica_cabecalho = f"Rubrica {rubrica_codigo_linha}"
                            if getattr(rubrica_calculo_obj, 'nome', None):
                                rubrica_cabecalho += f" - {rubrica_calculo_obj.nome}"

                            if valor_extrator == valor_esperado or diferenca_absoluta <= 0.01:
                                status = 'CORRETO'
                                justificativa = (
                                    f'{rubrica_cabecalho}. '
                                    f'{rubrica_descricao_calculo} | '
                                    f'Calculado por {origem_calculo}: {expressao_calculo}. '
                                    f'Usou valor_vencimento={valor_vencimento}, frequencia={frequencia}, carga_horaria={carga_horaria} → '
                                    f'Calculado = R$ {valor_esperado} ✓'
                                )
                                corretos += 1
                            elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                                status = 'VERIFICAR (Valores muito próximos)'
                                justificativa = (
                                    f'{rubrica_cabecalho}. '
                                    f'{rubrica_descricao_calculo} | '
                                    f'Calculado por {origem_calculo}: {expressao_calculo}. '
                                    f'Esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator} '
                                    f'(diferença: {diferenca_percentual}%)'
                                )
                                verificar += 1
                            else:
                                status = 'INCORRETO'
                                justificativa = (
                                    f'{rubrica_cabecalho}. '
                                    f'{rubrica_descricao_calculo} | '
                                    f'Calculado por {origem_calculo}: {expressao_calculo}. '
                                    f'Esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator}.'
                                )
                                incorretos += 1
                        else:
                            # Quando há rubrica customizada mas ela não produz valor esperado válido,
                            # a verificação segue para as regras especiais e comparações normais.
                            pass
                    elif rubrica_codigo_linha == '10502':
                        if frequencia is not None:
                            valor_esperado = round(valor_vencimento * 0.01 * frequencia, 2)
                            origem_calculo = 'Regra 10502'
                            expressao_calculo = '(valor_vencimento * 1%) * frequencia'
                        else:
                            valor_esperado = None
                            status = 'INCORRETO'
                            justificativa = 'Rubrica 10502: frequência não encontrada'
                            incorretos += 1
                    elif rubrica_codigo_linha in {'10020', '10024', '10025'}:
                        valor_esperado = _calcular_valor_esperado_rubrica_10020(
                            valor_vencimento,
                            frequencia,
                            valor_extrator,
                        )
                        origem_calculo = f'Regra {rubrica_codigo_linha}'
                        expressao_calculo = '((valor_vencimento / 30) * frequencia) * 0.25'
                    if valor_esperado is not None:
                        diferenca_absoluta = abs(valor_extrator - valor_esperado)
                        diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                        diferenca_percentual = round(diferenca_percentual, 2)
                        diferenca_absoluta = round(diferenca_absoluta, 2)

                        rubrica_cabecalho = f"Rubrica {rubrica_codigo_linha}"
                        if getattr(rubrica_calculo_obj, 'nome', None):
                            rubrica_cabecalho += f" - {rubrica_calculo_obj.nome}"

                        if valor_extrator == valor_esperado or diferenca_absoluta <= 0.01:
                            status = 'CORRETO'
                            justificativa = (
                                f'{rubrica_cabecalho}. '
                                f'{rubrica_descricao_calculo} | '
                                f'Calculado por {origem_calculo}: {expressao_calculo}. '
                                f'Usou valor_vencimento={valor_vencimento}, frequencia={frequencia}, carga_horaria={carga_horaria} → '
                                f'Calculado = R$ {valor_esperado} ✓'
                            )
                            corretos += 1
                        elif diferenca_percentual < 0.5:
                            status = 'VERIFICAR (Valores muito próximos)'
                            justificativa = (
                                f'{rubrica_cabecalho}. '
                                f'{rubrica_descricao_calculo} | '
                                f'Calculado por {origem_calculo}: {expressao_calculo}. '
                                f'Esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator} '
                                f'(diferença: {diferenca_percentual}%)'
                            )
                            verificar += 1
                        else:
                            status = 'INCORRETO'
                            justificativa = (
                                f'{rubrica_cabecalho}. '
                                f'{rubrica_descricao_calculo} | '
                                f'Calculado por {origem_calculo}: {expressao_calculo}. '
                                f'Esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator}.'
                            )
                            incorretos += 1

                if justificativa is None:
                    # REGRA ESPECIAL PARA RUBRICAS CALCULADAS PELO GRAU DE INSTRUÇÃO
                    # O valor do extrator deve ser igual a:
                    # valor_vencimento × percentual definido pelo grau de instrução
                    if rubrica_codigo_linha in rubricas_grau_instrucao and col_grau_instrucao:
                        if grau_instrucao_percentual is not None:
                            valor_esperado = round(valor_vencimento * grau_instrucao_percentual, 2)
                            diferenca_absoluta = abs(valor_extrator - valor_esperado)
                            diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                            diferenca_percentual = round(diferenca_percentual, 2)
                            diferenca_absoluta = round(diferenca_absoluta, 2)
                            
                            debug_print(f"\n[RUBRICA {rubrica_codigo_linha}] {nome_servidor}:")
                            debug_print(f"  Valor Vencimento: {valor_vencimento}")
                            debug_print(f"  Grau Instrução: {grau_instrucao}")
                            debug_print(f"  Percentual aplicado: {grau_instrucao_percentual}")
                            debug_print(f"  Valor Esperado: {valor_vencimento} × {grau_instrucao_percentual} = {valor_esperado}")
                            debug_print(f"  Valor Extrator: {valor_extrator}")
                            debug_print(f"  Diferença: {diferenca_absoluta} ({diferenca_percentual}%)")
                            
                            # Compara valores
                            if valor_extrator == valor_esperado:
                                status = 'CORRETO'
                                justificativa = f'Rubrica {rubrica_codigo_linha}: Grau de instrução "{grau_instrucao}" aplica {grau_instrucao_percentual * 100:.0f}% sobre R$ {valor_vencimento} = R$ {valor_esperado} ✓'
                                corretos += 1
                            elif diferenca_absoluta is not None and diferenca_absoluta <= 0.01:
                                # Tolerância para arredondamento: até R$ 0.01 de diferença é considerado CORRETO
                                status = 'CORRETO'
                                justificativa = f'Rubrica {rubrica_codigo_linha}: Grau de instrução "{grau_instrucao}" aplica {grau_instrucao_percentual * 100:.0f}% sobre R$ {valor_vencimento} = R$ {valor_esperado} (diferença de R$ {diferenca_absoluta} dentro da tolerância de arredondamento) ✓'
                                corretos += 1
                            elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                                status = 'VERIFICAR (Valores muito próximos)'
                                justificativa = f'Rubrica {rubrica_codigo_linha}: Grau de instrução "{grau_instrucao}" aplica {grau_instrucao_percentual * 100:.0f}% sobre R$ {valor_vencimento} = R$ {valor_esperado}, mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%)'
                                verificar += 1
                            else:
                                status = 'INCORRETO'
                                justificativa = f'Rubrica {rubrica_codigo_linha}: Grau de instrução "{grau_instrucao}" aplica {grau_instrucao_percentual * 100:.0f}% sobre R$ {valor_vencimento} = R$ {valor_esperado}, mas recebeu R$ {valor_extrator}'
                                incorretos += 1
                        else:
                            # Se não conseguiu obter grau de instrução, marca como incorreto
                            status = 'INCORRETO'
                            justificativa = f'Rubrica {rubrica_codigo_linha}: Grau de instrução não encontrado ou não reconhecido'
                            incorretos += 1
                            diferenca_absoluta = None
                            diferenca_percentual = None

                    elif str(rubrica).strip() == '10029':
                        if frequencia_percentual is not None:    
                            valor_esperado = round(valor_vencimento * 0.30, 2)
                            diferenca_absoluta = abs(valor_extrator - valor_esperado)
                            diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                            diferenca_percentual = round(diferenca_percentual, 2)
                            diferenca_absoluta = round(diferenca_absoluta, 2)

                            debug_print(f"\n[RUBRICA 10029] {nome_servidor}:")
                            debug_print(f"  Valor Vencimento: {valor_vencimento}")
                            debug_print(f"  Frequência original: {frequencia}")
                            debug_print(f"  Frequência percentual normalizada: {frequencia_percentual}")
                            debug_print(f"  Valor Esperado: {valor_vencimento} × {frequencia_percentual} = {valor_esperado}")
                            debug_print(f"  Valor Extrator: {valor_extrator}")
                            debug_print(f"  Diferença: {diferenca_absoluta} ({diferenca_percentual}%)")

                            # Compara valores
                            if valor_extrator == valor_esperado:
                                status = 'CORRETO'
                                justificativa = f'Rubrica 10029: R$ {valor_vencimento} × {frequencia_percentual} (frequência normalizada de {frequencia}) = R$ {valor_esperado} ✓'
                                corretos += 1
                            elif diferenca_absoluta is not None and diferenca_absoluta <= 0.01:
                                # Tolerância para arredondamento: até R$ 0.01 de diferença é considerado CORRETO
                                status = 'CORRETO'
                                justificativa = f'Rubrica 10029: R$ {valor_vencimento} × {frequencia_percentual} (frequência normalizada de {frequencia}) = R$ {valor_esperado} (diferença de R$ {diferenca_absoluta} dentro da tolerância de arredondamento) ✓'
                                corretos += 1
                            elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                                status = 'VERIFICAR (Valores muito próximos)'
                                justificativa = f'Rubrica 10029: R$ {valor_vencimento} × {frequencia_percentual} (frequência normalizada de {frequencia}) = R$ {valor_esperado}, mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%)'
                                verificar += 1
                            else:
                                status = 'INCORRETO'
                                justificativa = f'Rubrica 10029: Esperado R$ {valor_esperado} (R$ {valor_vencimento} × {frequencia_percentual} com frequência normalizada de {frequencia}), mas recebeu R$ {valor_extrator}'
                                incorretos += 1
                        else:
                            # Se não conseguiu obter frequência, marca como incorreto
                            status = 'INCORRETO'
                            justificativa = 'Rubrica 10029: Frequência não encontrada'
                            incorretos += 1
                            diferenca_absoluta = None
                            diferenca_percentual = None

                    elif str(rubrica).strip() == '10926':
                        if frequencia is not None:
                            frequencia_usada = min(frequencia, 22)
                            valor_esperado = round((640 / 22) * frequencia_usada, 2)
                            diferenca_absoluta = abs(valor_extrator - valor_esperado)
                            diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                            diferenca_percentual = round(diferenca_percentual, 2)
                            diferenca_absoluta = round(diferenca_absoluta, 2)

                            debug_print(f"\n[RUBRICA 10926] {nome_servidor}:")
                            debug_print(f"  Frequência original: {frequencia}")
                            debug_print(f"  Frequência usada: {frequencia_usada}")
                            debug_print(f"  Valor Esperado: (640 / 22) × {frequencia_usada} = {valor_esperado}")
                            debug_print(f"  Valor Extrator: {valor_extrator}")
                            debug_print(f"  Diferença: {diferenca_absoluta} ({diferenca_percentual}%)")

                            if valor_extrator > 640:
                                status = 'INCORRETO'
                                if frequencia_usada == 22 and frequencia > 22:
                                    justificativa = (
                                        f'Rubrica 10926: valor recebido maior que R$ 640 (R$ {valor_extrator}) é incorreto. '
                                        f'Frequencia > 22 é tratada como 22.'
                                    )
                                else:
                                    justificativa = f'Rubrica 10926: valor recebido maior que R$ 640 (R$ {valor_extrator}) é incorreto.'
                                incorretos += 1
                            elif valor_extrator == valor_esperado or (diferenca_absoluta is not None and diferenca_absoluta <= 0.01):
                                status = 'CORRETO'
                                justificativa = (
                                    f'Rubrica 10926: R$ {valor_extrator} está correto para frequência {frequencia} '
                                    f'(usando {frequencia_usada} quando frequencia > 22) = R$ {valor_esperado} ✓'
                                )
                                corretos += 1
                            elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                                status = 'VERIFICAR (Valores muito próximos)'
                                justificativa = f'Rubrica 10926: esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%)'
                                verificar += 1
                            else:
                                status = 'INCORRETO'
                                justificativa = f'Rubrica 10926: esperado R$ {valor_esperado}, mas recebeu R$ {valor_extrator}.'
                                incorretos += 1
                        else:
                            status = 'INCORRETO'
                            justificativa = 'Rubrica 10926: frequência não encontrada'
                            incorretos += 1
                            diferenca_absoluta = None
                            diferenca_percentual = None

                    elif str(rubrica).strip() == '10038':
                        if frequencia_percentual is not None:    
                            valor_esperado = round(valor_vencimento * 0.15, 2)
                            diferenca_absoluta = abs(valor_extrator - valor_esperado)
                            diferenca_percentual = (diferenca_absoluta / valor_esperado * 100) if valor_esperado != 0 else 0
                            diferenca_percentual = round(diferenca_percentual, 2)
                            diferenca_absoluta = round(diferenca_absoluta, 2)

                            debug_print(f"\n[RUBRICA 10038] {nome_servidor}:")
                            debug_print(f"  Valor Vencimento: {valor_vencimento}")
                            debug_print(f"  Frequência original: {frequencia}")
                            debug_print(f"  Frequência percentual normalizada: {frequencia_percentual}")
                            debug_print(f"  Valor Esperado: {valor_vencimento} × {frequencia_percentual} = {valor_esperado}")
                            debug_print(f"  Valor Extrator: {valor_extrator}")
                            debug_print(f"  Diferença: {diferenca_absoluta} ({diferenca_percentual}%)")

                            # Compara valores
                            if valor_extrator == valor_esperado:
                                status = 'CORRETO'
                                justificativa = f'Rubrica 10038: R$ {valor_vencimento} × {frequencia_percentual} (frequência normalizada de {frequencia}) = R$ {valor_esperado} ✓'
                                corretos += 1
                            elif diferenca_absoluta is not None and diferenca_absoluta <= 0.01:
                                # Tolerância para arredondamento: até R$ 0.01 de diferença é considerado CORRETO
                                status = 'CORRETO'
                                justificativa = f'Rubrica 10038: R$ {valor_vencimento} × {frequencia_percentual} (frequência normalizada de {frequencia}) = R$ {valor_esperado} (diferença de R$ {diferenca_absoluta} dentro da tolerância de arredondamento) ✓'
                                corretos += 1
                            elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                                status = 'VERIFICAR (Valores muito próximos)'
                                justificativa = f'Rubrica 10038: R$ {valor_vencimento} × {frequencia_percentual} (frequência normalizada de {frequencia}) = R$ {valor_esperado}, mas recebeu R$ {valor_extrator} (diferença: {diferenca_percentual}%)'
                                verificar += 1
                            else:
                                status = 'INCORRETO'
                                justificativa = f'Rubrica 10038: Esperado R$ {valor_esperado} (R$ {valor_vencimento} × {frequencia_percentual} com frequência normalizada de {frequencia}), mas recebeu R$ {valor_extrator}'
                                incorretos += 1
                        else:
                            # Se não conseguiu obter frequência, marca como incorreto
                            status = 'INCORRETO'
                            justificativa = 'Rubrica 10038: Frequência não encontrada'
                            incorretos += 1
                            diferenca_absoluta = None
                            diferenca_percentual = None

                    else:
                        # Comparação normal para outras rubricas
                        diferenca_absoluta = abs(valor_extrator - valor_vencimento)
                        diferenca_percentual = (diferenca_absoluta / valor_vencimento * 100) if valor_vencimento != 0 else 0
                        diferenca_percentual = round(diferenca_percentual, 2)
                        diferenca_absoluta = round(diferenca_absoluta, 2)
                        
                        # Compara valores
                        if valor_extrator == valor_vencimento:
                            status = 'CORRETO'
                            justificativa = f'Valor confere com vencimento: R$ {valor_vencimento} ✓'
                            corretos += 1
                        elif diferenca_absoluta is not None and diferenca_absoluta <= 0.01:
                            # Tolerância para arredondamento: até R$ 0.01 de diferença é considerado CORRETO
                            status = 'CORRETO'
                            justificativa = f'Valor confere com vencimento (diferença de R$ {diferenca_absoluta} dentro da tolerância) ✓'
                            corretos += 1
                        elif diferenca_percentual is not None and diferenca_percentual < 0.5:
                            # Valores muito próximos (< 0.5% de diferença) - marcar para verificação
                            status = 'VERIFICAR (Valores muito próximos)'
                            justificativa = f'Diferença de R$ {diferenca_absoluta} ({diferenca_percentual}%) em relação ao esperado R$ {valor_vencimento}'
                            verificar += 1
                        else:
                            status = 'INCORRETO'
                            justificativa = f'Esperado R$ {valor_vencimento}, mas recebeu R$ {valor_extrator}'
                            incorretos += 1

                if alternate_carga_match and status != 'CORRETO':
                    status = 'INCORRETO'
                    justificativa = (
                        f'Carga horária divergente: recebeu como '
                        f'{carga_recebida_texto or f"{alternate_carga_match["carga"]}h"} '
                        f'(R$ {valor_extrator}) em vez de {carga_horaria}h.'
                    )
            else:
                diferenca_absoluta = None
                diferenca_percentual = None
                status = 'INCORRETO'
                justificativa = 'Valores inválidos para comparação'
                incorretos += 1

        resultado = {
            'empresa': empresa,
            'dc_empresa': dc_empresa,
            'orgao': orgao,
            'rubrica': rubrica_codigo_linha,
            'cpf': cpf,
            'matricula': matricula,
            'nome_servidor': nome_servidor,
            'situacao_funcional': format_raw_value(row_extrator[col_situacao_funcional]) if col_situacao_funcional and col_situacao_funcional in df_extrator.columns else '',
            'status_servidor': format_raw_value(row_extrator[col_status]) if col_status and col_status in df_extrator.columns else '',
            'descricao_status': format_raw_value(row_extrator[col_descricao_status]) if col_descricao_status and col_descricao_status in df_extrator.columns else '',
            'cargo': format_raw_value(row_extrator[col_cargo]) if col_cargo and col_cargo in df_extrator.columns else '',
            'descricao_cargo': format_raw_value(row_extrator[col_descricao_cargo]) if col_descricao_cargo and col_descricao_cargo in df_extrator.columns else '',
            'data_admissao': format_raw_value(row_extrator[col_data_admissao]) if col_data_admissao and col_data_admissao in df_extrator.columns else '',
            'data_ingresso_ref_salarial': format_raw_value(row_extrator[col_data_ingresso_ref_salarial]) if col_data_ingresso_ref_salarial and col_data_ingresso_ref_salarial in df_extrator.columns else '',
            'data_afastamento': format_raw_value(row_extrator[col_data_afastamento]) if col_data_afastamento and col_data_afastamento in df_extrator.columns else '',
            'motivo_afastamento': format_raw_value(row_extrator[col_motivo_afastamento]) if col_motivo_afastamento and col_motivo_afastamento in df_extrator.columns else '',
            'motivo_desligamento': format_raw_value(row_extrator[col_motivo_desligamento]) if col_motivo_desligamento and col_motivo_desligamento in df_extrator.columns else '',
            'carga_horaria': format_raw_value(row_extrator[col_carga_horaria]) if col_carga_horaria and col_carga_horaria in df_extrator.columns else '',
            'carga_horaria_secundaria': format_raw_value(row_extrator[col_carga_horaria_secundaria]) if col_carga_horaria_secundaria and col_carga_horaria_secundaria in df_extrator.columns else '',
            'carga_horaria_total': row_extrator['carga_horaria_total'] if 'carga_horaria_total' in row_extrator else None,
            'ref_vertical': ref_v_raw if rubrica_eh_10014 or rubrica_eh_11187 else ref_v,
            'ref_horizontal': ref_h_raw if rubrica_eh_10014 or rubrica_eh_11187 else ref_h,
            'prov_desc': format_raw_value(row_extrator[col_prov_desc]) if col_prov_desc and col_prov_desc in df_extrator.columns else '',
            'valor_mes_anterior': norm_num(row_extrator[col_valor_mes_anterior]) if col_valor_mes_anterior and col_valor_mes_anterior in df_extrator.columns else None,
            'valor_mes_atual': norm_num(row_extrator[col_valor_mes_atual]) if col_valor_mes_atual and col_valor_mes_atual in df_extrator.columns else None,
            'valor_vencimento': valor_vencimento,
            'valor_total_recebido': valor_extrator,
            'frequencia': frequencia,
            'valor_calculado': (
                valor_esperado if valor_esperado is not None else 
                valor_vencimento if valor_vencimento is not None else 
                (alternate_carga_match['valor'] if alternate_carga_match else None)
            ),
            'status': status,
            'diferenca_absoluta': diferenca_absoluta,
            'diferenca_percentual': diferenca_percentual,
            'justificativa': justificativa,
        }
        resultados.append(resultado)
        if diferenca_percentual is not None:
            debug_print(f"{nome_servidor}: {status}")
            debug_print(f"  → {justificativa}")
        else:
            debug_print(f"{nome_servidor}: {status}")
            debug_print(f"  → {justificativa}")

    debug_print(f"\n=== RESUMO ===")
    debug_print(f"Total: {len(resultados)}")
    debug_print(f"Corretos: {corretos}")
    debug_print(f"Verificar: {verificar}")
    debug_print(f"Incorretos: {incorretos}")
    debug_print(f"=== FIM ===\n")

    return {
        'sucesso': True,
        'resultados': resultados,
        'total': len(resultados),
        'corretos': corretos,
        'verificar': verificar,
        'incorretos': incorretos,
        'mes_referencia': mes_referencia,
    }
