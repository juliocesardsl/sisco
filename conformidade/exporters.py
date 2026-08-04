import csv
from django.http import HttpResponse
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _calcular_mes_anterior(referencia_atual):
    """
    Calcula a referência do mês anterior baseado na referência atual.
    
    Aceita formatos:
    - "YYYY-MM" (ex: "2026-04") -> retorna "2026-03"
    - "MMYYYY" (ex: "042026") -> retorna "032026"
    - "" (vazio) -> retorna ""
    
    Args:
        referencia_atual: String com referência atual
        
    Returns:
        String com referência do mês anterior no mesmo formato
    """
    if not referencia_atual:
        return ""
    
    # Detecta formato e calcula mês anterior
    referencia_atual = str(referencia_atual).strip()
    
    # Formato YYYY-MM
    if '-' in referencia_atual and len(referencia_atual) == 7:
        try:
            ano, mes = referencia_atual.split('-')
            data = datetime.strptime(f"{ano}-{mes}-01", "%Y-%m-%d")
            data_anterior = data - relativedelta(months=1)
            return data_anterior.strftime("%Y-%m")
        except (ValueError, AttributeError):
            return referencia_atual
    
    # Formato MMYYYY
    elif len(referencia_atual) == 6 and referencia_atual.isdigit():
        try:
            mes = int(referencia_atual[:2])
            ano = int(referencia_atual[2:])
            data = datetime.strptime(f"{mes:02d}{ano}", "%m%Y")
            data_anterior = data - relativedelta(months=1)
            return f"{data_anterior.month:02d}{data_anterior.year}"
        except (ValueError, AttributeError):
            return referencia_atual
    
    # Retorna a referência atual se não conseguir processar
    return referencia_atual


def _calcular_versao_anterior(versao_atual):
    """
    Calcula a versão do mês anterior baseado na versão atual.
    
    Se versao_atual = "02", retorna "01"
    Se versao_atual = "1", retorna "0" (ou "1" se já for "1" ou "01")
    
    Args:
        versao_atual: String com versão atual
        
    Returns:
        String com versão do mês anterior
    """
    if not versao_atual:
        return ""
    
    versao_atual = str(versao_atual).strip()
    
    try:
        versao_num = int(versao_atual)
        if versao_num > 1:
            versao_anterior = versao_num - 1
            # Mantém o formato com zeros à esquerda (ex: "02" permanece "01")
            if len(versao_atual) > 1:
                return f"{versao_anterior:02d}"
            else:
                return str(versao_anterior)
        else:
            # Se é a primeira versão, não tem anterior
            return versao_atual
    except (ValueError, AttributeError):
        return versao_atual


def exportar_verificacoes_csv(queryset, status='todos'):
    """
    Exporta as verificações de conformidade para um arquivo CSV.

    Args:
        queryset: QuerySet com as verificações a exportar
        status: String indicando qual status foi filtrado

    Returns:
        HttpResponse com arquivo CSV para download
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="verificacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Rubrica', 'Empresa', 'Ano', 'Valor Mínimo', 'Valor Máximo', 'Valor Pago', 'Status', 'Data Verificação', 'Observações'])

    for verificacao in queryset:
        writer.writerow([
            verificacao.padrao.rubrica.nome,
            verificacao.padrao.empresa.nome,
            verificacao.padrao.ano,
            f"R$ {verificacao.padrao.valor_minimo:.2f}",
            f"R$ {verificacao.padrao.valor_maximo:.2f}",
            f"R$ {verificacao.valor_pago:.2f}",
            verificacao.get_status_display(),
            verificacao.data_verificacao.strftime('%d/%m/%Y %H:%M'),
            verificacao.observacoes,
        ])

    return response


def exportar_comparacao_excel(comparacao_data, rubrica_codigo, rubrica_nome):
    """
    Exporta a comparação mensal de rubricas para um arquivo Excel.

    Args:
        comparacao_data: Lista de dicts com dados da comparação
        rubrica_codigo: Código da rubrica
        rubrica_nome: Nome da rubrica

    Returns:
        HttpResponse com arquivo Excel para download
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparação"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    title_font = Font(bold=True, size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"Comparativo Mensal de Rubricas - {rubrica_codigo} ({rubrica_nome})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 25

    # Subtítulo com data
    ws.merge_cells('A2:M2')
    date_cell = ws['A2']
    date_cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}"
    date_cell.font = Font(italic=True, size=10, color="666666")
    date_cell.alignment = center_alignment

    # Linha em branco
    ws.append([])

    # Cabeçalhos do comparativo mensal
    headers = [
        'Empresa',
        'CPF',
        'Matrícula',
        'Nome',
        'CARGO',
        'RUBRICA',
        'DESCRIÇÃO DA RUBRICA',
        'Referência Anterior',
        'Versão Mês Anterior',
        'Valor Mês Anterior',
        'Referência Atual',
        'Versão Mês Atual',
        'Valor Mês Atual',
        'Variação (Dif.)',
        'Variação (%)',
        'Status da Variação',
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Dados
    row_num = 5
    total_anterior = 0
    total_atual = 0
    

    for item in comparacao_data:
        # Calcular referência anterior e versão anterior se não forem fornecidas
        referencia_anterior = item.get('referencia_anterior', '')
        if not referencia_anterior and item.get('referencia_atual'):
            referencia_anterior = _calcular_mes_anterior(item.get('referencia_atual', ''))
        
        # Versão anterior: se não fornecida, assume "01" (primeira versão/arquivo anterior)
        versao_anterior = item.get('versao_anterior', '')
        if not versao_anterior:
            versao_anterior = '01'
        
        # Versão atual: se não fornecida, assume "02" (segunda versão/arquivo atual)
        versao_atual = item.get('versao_atual', '')
        if not versao_atual:
            versao_atual = '02'
        
        ws.cell(row=row_num, column=1).value = item.get('empresa', '')
        ws.cell(row=row_num, column=2).value = item.get('cpf', '')
        ws.cell(row=row_num, column=3).value = item.get('matricula', '')
        ws.cell(row=row_num, column=4).value = item.get('nome_servidor', '')
        ws.cell(row=row_num, column=5).value = item.get('descricao_cargo', '')
        ws.cell(row=row_num, column=6).value = item.get('rubrica', '')
        ws.cell(row=row_num, column=7).value = item.get('dc_rubrica', '')
        ws.cell(row=row_num, column=8).value = referencia_anterior
        ws.cell(row=row_num, column=9).value = versao_anterior
        ws.cell(row=row_num, column=10).value = float(item.get('valor_anterior', 0))
        ws.cell(row=row_num, column=11).value = item.get('referencia_atual', '')
        ws.cell(row=row_num, column=12).value = versao_atual
        ws.cell(row=row_num, column=13).value = float(item.get('valor_atual', 0))
        ws.cell(row=row_num, column=14).value = float(item.get('diferenca', 0))
        variacao_pct = item.get('variacao_pct')
        if variacao_pct is not None:
            ws.cell(row=row_num, column=15).value = float(variacao_pct)
        ws.cell(row=row_num, column=16).value = item.get('status_variacao') or ''

        total_anterior += float(item.get('valor_anterior', 0))
        total_atual += float(item.get('valor_atual', 0))
        
        # Aplicar estilos e cores
        for col in range(1, 17):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border

            if col in (10, 13, 14, 15):
                cell.alignment = right_alignment
                if col == 14:
                    diferenca = float(item.get('diferenca', 0))
                    if diferenca > 0:
                        cell.font = Font(bold=True, color="28A745")
                    elif diferenca < 0:
                        cell.font = Font(bold=True, color="DC3545")
                elif col == 15:
                    variacao_pct = item.get('variacao_pct')
                    if variacao_pct is not None:
                        if float(variacao_pct) > 0:
                            cell.font = Font(bold=True, color="28A745")
                        elif float(variacao_pct) < 0:
                            cell.font = Font(bold=True, color="DC3545")
            else:
                cell.alignment = left_alignment
        
        row_num += 1

    # Linha de totais
    total_row = row_num
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=10).value = total_anterior
    ws.cell(row=total_row, column=13).value = total_atual

    if total_anterior > 0:
        total_variacao_pct = ((total_atual - total_anterior) / total_anterior) * 100
        ws.cell(row=total_row, column=15).value = total_variacao_pct

    for col in range(1, 17):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.border = thin_border
        if col in (11, 12, 13, 14):
            cell.alignment = right_alignment
        else:
            cell.alignment = left_alignment

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 36
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18
    ws.column_dimensions['M'].width = 16
    ws.column_dimensions['N'].width = 14
    ws.column_dimensions['O'].width = 14
    ws.column_dimensions['P'].width = 28

    # Formatar células numéricas como currency
    for row in ws.iter_rows(min_row=5, max_row=total_row, min_col=10, max_col=13):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'
    
    for row in ws.iter_rows(min_row=5, max_row=total_row, min_col=15, max_col=15):
        for cell in row:
            cell.number_format = '0.00"%"'

    # Congelar linha de cabeçalho
    ws.freeze_panes = "A5"
    
    # Aplicar AutoFilter para poder filtrar cada coluna
    ws.auto_filter.ref = f'A4:P{total_row}'

    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="comparacao_rubricas_{rubrica_codigo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response
