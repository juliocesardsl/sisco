from io import BytesIO
from pathlib import Path

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase

from conformidade.models import Rubrica, Empresa, PadraoConformidade, VerificacaoConformidade

from conformidade.agent import gerar_resposta_agente
from conformidade.exporters import exportar_comparacao_excel
from conformidade.verificacao_utils import (
    calcular_status_variacao,
    comparar_extrator_por_mes,
    determinar_status_comparacao,
    _mapear_grau_instrucao_para_percentual,
)


class ComparacaoMensalTests(SimpleTestCase):
    def test_determinar_status_comparacao_retorna_status_novo_e_removido(self):
        self.assertEqual(determinar_status_comparacao(0, 100), 'novo')
        self.assertEqual(determinar_status_comparacao(100, 0), 'removido')
        self.assertEqual(determinar_status_comparacao(100, 150), 'aumento')
        self.assertEqual(determinar_status_comparacao(150, 100), 'reducao')
        self.assertEqual(determinar_status_comparacao(100, 100), 'sem-variacao')

    def test_calcular_status_variacao_retorna_rotulos_esperados(self):
        self.assertEqual(calcular_status_variacao(0, 100), 'Sem pagamento no mês anterior')
        self.assertEqual(calcular_status_variacao(100, 0), 'Sem pagamento no mês atual')
        self.assertEqual(calcular_status_variacao(100, 150), 'Houve aumento')
        self.assertEqual(calcular_status_variacao(150, 100), 'Houve redução')
        self.assertEqual(calcular_status_variacao(100, 100), 'Não houve variação')
        self.assertEqual(calcular_status_variacao(0, 0), '')


    def test_comparacao_mesal_identifica_rubrica_com_coluna_rubrica(self):
        base_dir = Path(__file__).resolve().parent.parent
        arquivo_anterior = base_dir / 'Arquivos' / 'Pagamento_10004_032026.xlsx'
        arquivo_atual = base_dir / 'Arquivos' / 'Pagamento_10004_042026.xlsx'

        anterior = SimpleUploadedFile(
            arquivo_anterior.name,
            arquivo_anterior.read_bytes(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        atual = SimpleUploadedFile(
            arquivo_atual.name,
            arquivo_atual.read_bytes(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        resultado = comparar_extrator_por_mes(anterior, atual, '10004')

        self.assertNotIn('erro', resultado)
        self.assertGreater(len(resultado['comparacao']), 0)
        self.assertTrue(any(item.get('referencia_atual') for item in resultado['comparacao']))

        primeiro_item = resultado['comparacao'][0]
        self.assertIn('rubrica', primeiro_item)
        self.assertIn('dc_rubrica', primeiro_item)
        self.assertIn('dc_empresa', primeiro_item)
        self.assertTrue(primeiro_item.get('rubrica'))
        self.assertTrue(primeiro_item.get('dc_rubrica'))

    def test_comparacao_mesal_usa_codigo_rubrica_real_quando_coluna_rubrica_tem_valores_monetarios(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['EMPRESA', 'MATRICULA', 'RUBRICA', 'ID_PROV_DESC', 'DC_RUBRICA', 'VL_RUBRICA'])
        ws.append(['004', '01269267', '10755.52', '10004', 'Rubrica Teste', '7729.23'])
        ws.append(['004', '01269267', '10755.52', '10004', 'Rubrica Teste', '5812.14'])
        bytes_data = BytesIO()
        wb.save(bytes_data)
        content = bytes_data.getvalue()

        anterior = SimpleUploadedFile(
            'Pagamento_10004_032026.xlsx',
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        atual = SimpleUploadedFile(
            'Pagamento_10004_042026.xlsx',
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        resultado = comparar_extrator_por_mes(anterior, atual, '10004')

        self.assertNotIn('erro', resultado)
        self.assertGreater(len(resultado['comparacao']), 0)
        primeiro_item = resultado['comparacao'][0]
        self.assertEqual(primeiro_item.get('rubrica'), '10004')
        self.assertEqual(primeiro_item.get('dc_rubrica'), 'Rubrica Teste')

    def test_exportar_comparacao_excel_inclui_coluna_cargo(self):
        comparacao = [
            {
                'empresa': '001',
                'cpf': '12345678901',
                'matricula': '0001',
                'nome_servidor': 'Servidor Teste',
                'descricao_cargo': 'ANALISTA',
                'dc_rubrica': 'Descrição da Rubrica',
                'referencia_anterior': '052026',
                'referencia_atual': '062026',
                'valor_anterior': 10.0,
                'valor_atual': 15.0,
                'diferenca': 5.0,
                'variacao_pct': 50.0,
                'status_variacao': 'Houve aumento',
            }
        ]

        response = exportar_comparacao_excel(comparacao, '10004', 'Teste Rubrica')

        self.assertEqual(response.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        ws = workbook.active

        headers = [cell.value for cell in ws[4]]
        self.assertEqual(headers, [
            'Empresa', 'CPF', 'Matrícula', 'Nome', 'CARGO', 'RUBRICA', 'DESCRIÇÃO DA RUBRICA',
            'Referência Anterior', 'Versão Mês Anterior', 'Valor Mês Anterior',
            'Referência Atual', 'Versão Mês Atual', 'Valor Mês Atual',
            'Variação (Dif.)', 'Variação (%)', 'Status da Variação'
        ])
        self.assertEqual(ws.cell(row=5, column=5).value, 'ANALISTA')
        self.assertEqual(ws.cell(row=5, column=8).value, '052026')
        self.assertEqual(ws.cell(row=5, column=9).value, '01')
        self.assertEqual(ws.cell(row=5, column=10).value, 10.0)
        self.assertEqual(ws.cell(row=5, column=11).value, '062026')
        self.assertEqual(ws.cell(row=5, column=12).value, '02')
        self.assertEqual(ws.cell(row=5, column=13).value, 15.0)


class RegraRubrica11033Tests(SimpleTestCase):
    def test_mapeia_percentuais_por_grau_de_instrucao(self):
        casos = {
            'ENSINO MEDIO COMPLETO': 0.09,
            '2ª GRADUAÇÃO': 0.09,
            'SEGUNDA GRADUAÇÃO': 0.09,
            'GRADUAÇÃO': 0.13,
            'ESPECIALIZAÇÃO': 0.20,
            'MESTRADO': 0.30,
            'DOUTORADO': 0.35,
        }

        for grau_instrucao, percentual in casos.items():
            with self.subTest(grau_instrucao=grau_instrucao):
                self.assertEqual(
                    _mapear_grau_instrucao_para_percentual(grau_instrucao, '11033'),
                    percentual,
                )

        self.assertEqual(
            _mapear_grau_instrucao_para_percentual('ENSINO MEDIO COMPLETO', '10582'),
            0.10,
        )

    def test_mapeia_regra_especifica_da_rubrica_11110(self):
        casos = {
            '2ª GRADUAÇÃO': 0.13,
            'SEGUNDA GRADUAÇÃO': 0.13,
            'ESPECIALIZAÇÃO': 0.20,
            'MESTRADO': 0.30,
            'DOUTORADO': 0.35,
        }

        for grau_instrucao, percentual in casos.items():
            with self.subTest(grau_instrucao=grau_instrucao):
                self.assertEqual(
                    _mapear_grau_instrucao_para_percentual(grau_instrucao, '11110'),
                    percentual,
                )

        self.assertIsNone(_mapear_grau_instrucao_para_percentual('GRADUAÇÃO', '11110'))
        self.assertIsNone(_mapear_grau_instrucao_para_percentual('ENSINO MEDIO COMPLETO', '11110'))

    def test_mapeia_regra_especifica_da_rubrica_11171(self):
        casos = {
            'GRADUAÇÃO': 0.15,
            'PÓS-GRADUAÇÃO': 0.15,
            'MESTRADO': 0.35,
            'DOUTORADO': 0.40,
        }

        for grau_instrucao, percentual in casos.items():
            with self.subTest(grau_instrucao=grau_instrucao):
                self.assertEqual(
                    _mapear_grau_instrucao_para_percentual(grau_instrucao, '11171'),
                    percentual,
                )

    def test_mapeia_regra_especifica_da_rubrica_11189(self):
        casos = {
            'GRADUAÇÃO': 0.15,
            '2ª GRADUAÇÃO': 0.15,
            'SEGUNDA GRADUAÇÃO': 0.15,
            'ESPECIALIZAÇÃO': 0.25,
            'MESTRADO': 0.35,
            'DOUTORADO': 0.40,
        }

        for grau_instrucao, percentual in casos.items():
            with self.subTest(grau_instrucao=grau_instrucao):
                self.assertEqual(
                    _mapear_grau_instrucao_para_percentual(grau_instrucao, '11189'),
                    percentual,
                )

        for grau_instrucao, percentual in casos.items():
            with self.subTest(rubrica='11190', grau_instrucao=grau_instrucao):
                self.assertEqual(
                    _mapear_grau_instrucao_para_percentual(grau_instrucao, '11190'),
                    percentual,
                )


class AgentAssistantTests(TestCase):
    def test_gerar_resposta_agente_resumo_geral_com_dados_reais(self):
        rubrica = Rubrica.objects.create(codigo='10004', nome='Rubrica Teste')
        empresa = Empresa.objects.create(nome='Empresa Teste', codigo='EMP001', cnpj='12.345.678/0001-99', razao_social='Empresa Teste Ltda')
        padrao = PadraoConformidade.objects.create(
            rubrica=rubrica,
            empresa=empresa,
            ano=2026,
            valor_minimo=100,
            valor_maximo=200,
            carga_horaria=40,
        )
        VerificacaoConformidade.objects.create(
            padrao=padrao,
            rubrica='10004',
            ano_referencia=2026,
            carga_horaria=40,
            valor_pago=150,
            status='correto',
        )

        resposta = gerar_resposta_agente('resumo geral do sistema')

        self.assertIn('Rubricas', resposta)
        self.assertIn('Empresas', resposta)
        self.assertIn('Padrões', resposta)
        self.assertIn('Verificações', resposta)
        self.assertIn('1', resposta)


class Rubrica10014Tests(TestCase):
    def test_processar_verificacao_para_rubrica_10014_usa_apenas_extrator(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF FUNCIONAL VERTICAL',
            'REF FUNCIONAL HORIZONTAL',
            'VALOR',
            'NOME'
        ])
        ws_ext.append(['10014', 2026, 40, 'CPC-08', 'CPC-08', 2940.0, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            None,
            arquivo_extrator,
            '10014',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10014')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 2940.0)
        self.assertIn('10014', item['justificativa'])

    def test_processar_verificacao_10014_aplica_frequencia_para_valor_esperado(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF FUNCIONAL VERTICAL',
            'REF FUNCIONAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10014', 2026, 40, 'CPC-04', 'CPC-04', 1134.0, 21, 'Sonia da Conceicao Silva'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator_freq.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            None,
            arquivo_extrator,
            '10014',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 1134.0)
        self.assertIn('CPC-04', item['justificativa'])

    def test_processar_verificacao_combina_ref_funcional_vertical_e_horizontal(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF FUNCIONAL VERTICAL',
            'REF FUNCIONAL HORIZONTAL',
            'VALOR',
            'NOME'
        ])
        ws_ext.append(['10014', 2026, 40, 'CPC', '03', 1450.0, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator_split.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument/spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            None,
            arquivo_extrator,
            '10014',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 1450.0)
        self.assertIn('CPC-03', item['justificativa'])


class Rubrica11187Tests(TestCase):
    def test_processar_verificacao_11187_compara_com_rubrica_10014_ja_validada(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF FUNCIONAL VERTICAL',
            'REF FUNCIONAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10014', 2026, 40, 'CPC-04', 'CPC-04', 1134.0, 21, 'Sonia da Conceicao Silva'])
        ws_ext.append(['11187', 2026, 40, 'CPC-04', 'CPC-04', 378.0, 21, 'Sonia da Conceicao Silva'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator_11187.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            None,
            arquivo_extrator,
            '11187',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 378.0)
        self.assertIn('11187', item['justificativa'])
        self.assertIn('10014', item['justificativa'])


class VerificacaoFormulaTests(TestCase):
    def test_processar_verificacao_carrega_rubrica_do_banco_quando_nao_recebe_objeto(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10502',
            codigo='10502',
            descricao='Rubrica carregada do banco',
            criterio_calculo_rubrica='valor_vencimento * frequencia / 100',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10502')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Critério de cálculo', item['justificativa'])

    def test_processar_verificacao_soma_carga_horaria_com_secundaria(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'DATA_VIGENCIA',
            'REFERENCIA',
            'ANO_REFER',
            'MES_REFER',
            'CARGA_HORARIA',
            'REFER_SALARIAL',
            'REFER_SALARIAL_VERTICAL',
            'REFER_SALARIAL_HORIZONTAL',
            'VL_VENCIMENTO',
            'CARREIRA',
            'CATEGORIA',
            'CLASSE',
            'PADRAO',
            'ORDEM',
            'PUBLICACAO'
        ])
        ws_venc.append(['2026-01-01', 'S1', 2026, 1, 40, 'S1', 'S1', 'H1', 300, 'X', 'Y', 'Z', 'P', '1', 'PUB'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'CARGA HORARIA SECUNDARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 20, 20, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['carga_horaria_total'], 40)
        self.assertEqual(item['carga_horaria'], '20')
        self.assertEqual(item['carga_horaria_secundaria'], '20')
        self.assertEqual(item['status'], 'CORRETO')

    def test_processar_verificacao_ignora_coluna_secundaria_quando_ela_veio_antes_da_principal(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'DATA_VIGENCIA',
            'REFERENCIA',
            'ANO_REFER',
            'MES_REFER',
            'CARGA_HORARIA',
            'REFER_SALARIAL',
            'REFER_SALARIAL_VERTICAL',
            'REFER_SALARIAL_HORIZONTAL',
            'VL_VENCIMENTO',
            'CARREIRA',
            'CATEGORIA',
            'CLASSE',
            'PADRAO',
            'ORDEM',
            'PUBLICACAO'
        ])
        ws_venc.append(['2026-01-01', 'S1', 2026, 1, 40, 'S1', 'S1', 'H1', 300, 'X', 'Y', 'Z', 'P', '1', 'PUB'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA SECUNDARIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 10, 30, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado_40 = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado_40)
        self.assertEqual(resultado_40['total'], 1)
        self.assertEqual(resultado_40['resultados'][0]['carga_horaria_total'], 40)
        self.assertEqual(resultado_40['resultados'][0]['carga_horaria'], '30')
        self.assertEqual(resultado_40['resultados'][0]['carga_horaria_secundaria'], '10')

    def test_processar_verificacao_agrega_carga_horaria_secundaria_no_bucket_correspondente(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'DATA_VIGENCIA',
            'REFERENCIA',
            'ANO_REFER',
            'MES_REFER',
            'CARGA_HORARIA',
            'REFER_SALARIAL',
            'REFER_SALARIAL_VERTICAL',
            'REFER_SALARIAL_HORIZONTAL',
            'VL_VENCIMENTO',
            'CARREIRA',
            'CATEGORIA',
            'CLASSE',
            'PADRAO',
            'ORDEM',
            'PUBLICACAO'
        ])
        ws_venc.append(['2026-01-01', 'S1', 2026, 1, 40, 'S1', 'S1', 'H1', 300, 'X', 'Y', 'Z', 'P', '1', 'PUB'])
        ws_venc.append(['2026-01-01', 'S1', 2026, 1, 30, 'S1', 'S1', 'H1', 250, 'X', 'Y', 'Z', 'P', '1', 'PUB'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'CARGA HORARIA SECUNDARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 30, 10, 'S1', 'H1', 150, 50, 'Servidor 40h'])
        ws_ext.append(['10502', 2026, 20, 10, 'S1', 'H1', 120, 50, 'Servidor 30h'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado_40 = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )
        self.assertNotIn('erro', resultado_40)
        self.assertEqual(resultado_40['total'], 1)
        self.assertEqual(resultado_40['resultados'][0]['nome_servidor'], 'Servidor 40h')
        self.assertEqual(resultado_40['resultados'][0]['carga_horaria_total'], 40)

        resultado_30 = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            30,
        )
        self.assertNotIn('erro', resultado_30)
        self.assertEqual(resultado_30['total'], 1)
        self.assertEqual(resultado_30['resultados'][0]['nome_servidor'], 'Servidor 30h')
        self.assertEqual(resultado_30['resultados'][0]['carga_horaria_total'], 30)

    def test_processar_verificacao_usa_carga_horaria_total_na_justificativa_quando_tem_secundaria(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 20, 150, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'CARGA HORARIA SECUNDARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 20, 20, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertIn('20h + 20h = 40h', item['justificativa'])
        self.assertNotIn('recebeu como 20h (R$', item['justificativa'])

    def test_processar_verificacao_seleciona_coluna_nome_servidor_correta(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME EMPRESA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Empresa Exemplo', 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['nome_servidor'], 'Servidor Teste')
        self.assertEqual(item['status'], 'CORRETO')

    def test_processar_verificacao_10502_aceita_colunas_vencimento_alternativas(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'DATA_VIGENCIA',
            'REFERENCIA',
            'ANO_REFER',
            'MES_REFER',
            'CARGA_HORARIA',
            'REFER_SALARIAL',
            'REFER_SALARIAL_VERTICAL',
            'REFER_SALARIAL_HORIZONTAL',
            'VL_VENCIMENTO',
            'CARREIRA',
            'CATEGORIA',
            'CLASSE',
            'PADRAO',
            'ORDEM',
            'PUBLICACAO'
        ])
        ws_venc.append(['2026-01-01', 'S1', 2026, 1, 40, 'S1', 'S1', 'H1', 300, 'X', 'Y', 'Z', 'P', '1', 'PUB'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Regra 10502', item['justificativa'])

    def test_processar_verificacao_10502_aplica_regra_vencimento_1pct_frequencia(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10502')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Regra 10502', item['justificativa'])

    def test_processar_verificacao_10502_prioriza_regra_especial_mesmo_com_formula_cadastrada(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10502',
            codigo='10502',
            descricao='Teste regra especial 10502',
            valor_padrao=0,
            formula_calculo='valor_vencimento * 0.5'
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Regra 10502', item['justificativa'])

    def test_processar_verificacao_usa_formula_da_rubrica(self):
        from conformidade.verificacao_utils import processar_verificacao

        rubrica = Rubrica.objects.create(
            nome='10502',
            codigo='10502',
            descricao='Teste fórmula rubrica',
            valor_padrao=0,
            formula_calculo='valor_vencimento * frequencia / 100'
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10502',
            2026,
            40,
            rubrica_obj=rubrica,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Regra de cálculo', item['justificativa'])

    def test_processar_verificacao_usa_criterio_da_rubrica_e_rubrica_da_linha(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10502',
            codigo='10502',
            descricao='Teste critério rubrica',
            valor='Aplicar o percentual informado sobre o valor do vencimento',
            criterio_calculo_rubrica='valor_vencimento * frequencia / 100',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10502'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10502')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Critério de cálculo', item['justificativa'])
        self.assertIn('Valor', item['justificativa'])

    def test_processar_verificacao_extrai_formula_de_criterio_textual(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='20001',
            codigo='20001',
            descricao='Teste critério textual',
            criterio_calculo_rubrica='Aplicar 50% sobre o valor_vencimento',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['20001', 2026, 40, 'S1', 'H1', 150, 1, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10502'),
        )

        self.assertNotIn('erro', resultado)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '20001')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 150.0)
        self.assertIn('Aplicar 50%', item['justificativa'])

    def test_processar_verificacao_10020_com_texto_legal_pos_formula(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10020',
            codigo='10020',
            descricao='Teste rubrica 10020',
            valor='vencimento x 25% Aplica-se o disposto nesta Lei',
            criterio_calculo_rubrica='valor_vencimento x 25% Aplica-se o disposto nesta Lei',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 7729.23, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME',
            'CPF'
        ])
        ws_ext.append(['10020', 2026, 40, 'S1', 'H1', 1932.31, 25, 'Servidor Teste', '11122233344'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10020'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10020')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 1932.31)
        self.assertIn('25%', item['justificativa'])

    def test_processar_verificacao_10024_usa_mesma_regra_da_10020(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10024',
            codigo='10024',
            descricao='Teste rubrica 10024 com frequência',
            valor='((valor_vencimento / 30) * frequencia) * 0.25',
            criterio_calculo_rubrica='((valor_vencimento / 30) * frequencia) * 0.25',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 10762.32, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME',
            'CPF'
        ])
        ws_ext.append(['10024', 2026, 40, 'S1', 'H1', 1883.41, 21, 'Servidor Teste', '11122233344'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10024'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10024')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 1883.41)
        self.assertIn('Regra 10024', item['justificativa'])

    def test_processar_verificacao_10020_usa_frequencia(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10020',
            codigo='10020',
            descricao='Teste rubrica 10020 com frequência',
            valor='vencimento x 25%',
            criterio_calculo_rubrica='valor_vencimento * 0.25',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 10762.32, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME',
            'CPF'
        ])
        ws_ext.append(['10020', 2026, 40, 'S1', 'H1', 1883.40, 21, 'Servidor Teste', '11122233344'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10020'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10020')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 1883.4)
        self.assertIn('frequencia', item['justificativa'])

    def test_processar_verificacao_10502_aceita_frequencia_decimal(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10502',
            codigo='10502',
            descricao='Teste rubrica 10502',
            valor='valor do vencimento x porcentagem da frequencia',
            criterio_calculo_rubrica='valor_vencimento * frequencia_percentual',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 7729.23, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME',
            'CPF'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 2782.52, 0.36, 'Servidor Teste', '15210995372'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10502'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10502')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 2782.52)
        self.assertIn('frequencia_percentual', item['justificativa'])

    def test_processar_verificacao_10008_apenas_aposentado_usa_valor_vencimento(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10008',
            codigo='10008',
            descricao='Teste rubrica 10008 aposentado',
            criterio_calculo_rubrica='valor_vencimento * 0.5',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 1000, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'DC_SITUACAO_FUNCIONAL',
            'NOME'
        ])
        ws_ext.append(['10008', 2026, 40, 'S1', 'H1', 1000, 'APOSENTADO ESTATUTARIO', 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10008',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10008'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10008')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 1000.0)
        self.assertIn('Tabela de Vencimento', item['justificativa'])

    def test_processar_verificacao_10008_nao_aposentado_usa_formula_da_rubrica(self):
        from conformidade.verificacao_utils import processar_verificacao

        Rubrica.objects.create(
            nome='10008',
            codigo='10008',
            descricao='Teste rubrica 10008 não aposentado',
            criterio_calculo_rubrica='valor_vencimento * 0.5',
            base_calculo='valor_vencimento',
        )

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 1000, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'DC_STATUS',
            'NOME'
        ])
        ws_ext.append(['10008', 2026, 40, 'S1', 'H1', 500, 'NORMAL', 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10008',
            2026,
            40,
            rubrica_obj=Rubrica.objects.get(codigo='10008'),
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['rubrica'], '10008')
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 500.0)
        self.assertIn('Calculado por', item['justificativa'])

    def test_processar_verificacao_10582_usa_grau_de_instrucao(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 5796.91, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'DC_GRAU_INSTRUCAO',
            'NOME'
        ])
        ws_ext.append(['10582', 2026, 40, 'S1', 'H1', 579.69, 'ENSINO MEDIO COMPLETO', 'Servidor 1'])
        ws_ext.append(['10582', 2026, 40, 'S1', 'H1', 869.54, 'ENSINO SUPERIOR COMPLETO', 'Servidor 2'])
        ws_ext.append(['10582', 2026, 40, 'S1', 'H1', 1449.23, 'ESPECIALIZACAO', 'Servidor 3'])
        ws_ext.append(['10582', 2026, 40, 'S1', 'H1', 2028.92, 'MESTRADO', 'Servidor 4'])
        ws_ext.append(['10582', 2026, 40, 'S1', 'H1', 2318.76, 'DOUTORADO', 'Servidor 5'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'Pagamento_sem_552_652_052026_grau_instrucao.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10582',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 5)
        self.assertEqual(resultado['corretos'], 5)

        esperado_por_nome = {
            'Servidor 1': 579.69,
            'Servidor 2': 869.54,
            'Servidor 3': 1449.23,
            'Servidor 4': 2028.92,
            'Servidor 5': 2318.76,
        }
        for item in resultado['resultados']:
            self.assertEqual(item['status'], 'CORRETO')
            self.assertEqual(item['valor_calculado'], esperado_por_nome[item['nome_servidor']])
            self.assertIn('Grau de instrução', item['justificativa'])

    def test_processar_verificacao_10584_usa_grau_de_instrucao(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 5796.91, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento_10584.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'DC_GRAU_INSTRUCAO',
            'NOME'
        ])
        ws_ext.append(['10584', 2026, 40, 'S1', 'H1', 579.69, 'ENSINO MEDIO COMPLETO', 'Servidor 10584'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'Pagamento_10584_grau_instrucao.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10584',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        self.assertEqual(resultado['corretos'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 579.69)
        self.assertIn('Grau de instrução', item['justificativa'])

    def test_processar_verificacao_10512_usa_regra_da_10582(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 5796.91])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento_10512.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'DC_GRAU_INSTRUCAO',
            'NOME',
        ])
        ws_ext.append(['10512', 2026, 40, 'S1', 'H1', 869.54, 'ENSINO SUPERIOR COMPLETO', 'Servidor 10512'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator_10512.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(arquivo_vencimento, arquivo_extrator, '10512', 2026, 40)

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        self.assertEqual(resultado['corretos'], 1)
        item = resultado['resultados'][0]
        self.assertEqual(item['status'], 'CORRETO')
        self.assertEqual(item['valor_calculado'], 869.54)
        self.assertIn('Grau de instrução', item['justificativa'])

    def test_processar_verificacao_10926_aplica_regra_640_22_frequencia(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 1000, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10926', 2026, 40, 'S1', 'H1', 640, 22, 'Servidor Correto'])
        ws_ext.append(['10926', 2026, 40, 'S1', 'H1', 650, 22, 'Servidor Incorreto'])
        ws_ext.append(['10926', 2026, 40, 'S1', 'H1', 640, 25, 'Servidor Frequencia 25'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'Pagamento_10926.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10926',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 3)
        self.assertEqual(resultado['corretos'], 2)
        self.assertEqual(resultado['incorretos'], 1)

        item_correto = next(item for item in resultado['resultados'] if item['nome_servidor'] == 'Servidor Correto')
        item_incorreto = next(item for item in resultado['resultados'] if item['nome_servidor'] == 'Servidor Incorreto')
        item_freq_25 = next(item for item in resultado['resultados'] if item['nome_servidor'] == 'Servidor Frequencia 25')

        self.assertEqual(item_correto['status'], 'CORRETO')
        self.assertEqual(item_correto['valor_calculado'], 640.0)
        self.assertIn('Rubrica 10926', item_correto['justificativa'])

        self.assertEqual(item_incorreto['status'], 'INCORRETO')
        self.assertIn('maior que R$ 640', item_incorreto['justificativa'])

        self.assertEqual(item_freq_25['status'], 'CORRETO')
        self.assertEqual(item_freq_25['valor_calculado'], 640.0)
        self.assertIn('frequencia > 22', item_freq_25['justificativa'])

    def test_processar_verificacao_10605_usa_grau_de_instrucao(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 5796.91, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'DC_GRAU_INSTRUCAO',
            'NOME'
        ])
        ws_ext.append(['10605', 2026, 40, 'S1', 'H1', 579.69, 'ENSINO MEDIO COMPLETO', 'Servidor 1'])
        ws_ext.append(['10605', 2026, 40, 'S1', 'H1', 869.54, 'ENSINO SUPERIOR COMPLETO', 'Servidor 2'])
        ws_ext.append(['10605', 2026, 40, 'S1', 'H1', 1449.23, 'ESPECIALIZACAO', 'Servidor 3'])
        ws_ext.append(['10605', 2026, 40, 'S1', 'H1', 2028.92, 'MESTRADO', 'Servidor 4'])
        ws_ext.append(['10605', 2026, 40, 'S1', 'H1', 2318.76, 'DOUTORADO', 'Servidor 5'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'Pagamento_sem_552_652_052026_grau_instrucao_10605.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '10605',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 5)
        self.assertEqual(resultado['corretos'], 5)

        esperado_por_nome = {
            'Servidor 1': 579.69,
            'Servidor 2': 869.54,
            'Servidor 3': 1449.23,
            'Servidor 4': 2028.92,
            'Servidor 5': 2318.76,
        }
        for item in resultado['resultados']:
            self.assertEqual(item['status'], 'CORRETO')
            self.assertEqual(item['valor_calculado'], esperado_por_nome[item['nome_servidor']])
            self.assertIn('Grau de instrução', item['justificativa'])

    def test_processar_verificacao_sem_rubrica_faz_todas_as_rubricas(self):
        from conformidade.verificacao_utils import processar_verificacao

        wb_venc = openpyxl.Workbook()
        ws_venc = wb_venc.active
        ws_venc.append([
            'REFERENCIA DE VENCIMENTO VERTICAL',
            'REFERENCIA DE VENCIMENTO HORIZONTAL',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'VALOR',
            'FILTRO VENCIMENTO'
        ])
        ws_venc.append(['S1', 'H1', 2026, 40, 300, 'Rubrica ligada ao vencimento?'])
        bytes_venc = BytesIO()
        wb_venc.save(bytes_venc)
        bytes_venc.seek(0)
        arquivo_vencimento = SimpleUploadedFile(
            'vencimento.xlsx',
            bytes_venc.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb_ext = openpyxl.Workbook()
        ws_ext = wb_ext.active
        ws_ext.append([
            'PROV/DESC',
            'ANO REFERENCIA',
            'CARGA HORARIA',
            'REF SALARIAL VERTICAL',
            'REF SALARIAL HORIZONTAL',
            'VALOR',
            'FREQUENCIA',
            'NOME'
        ])
        ws_ext.append(['10502', 2026, 40, 'S1', 'H1', 150, 50, 'Servidor Teste'])
        bytes_ext = BytesIO()
        wb_ext.save(bytes_ext)
        bytes_ext.seek(0)
        arquivo_extrator = SimpleUploadedFile(
            'extrator.xlsx',
            bytes_ext.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        resultado = processar_verificacao(
            arquivo_vencimento,
            arquivo_extrator,
            '',
            2026,
            40,
        )

        self.assertNotIn('erro', resultado)
        self.assertEqual(resultado['total'], 1)
        self.assertEqual(resultado['incorretos'], 1)
